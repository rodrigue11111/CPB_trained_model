# -*- coding: utf-8 -*-
"""Analyse interpretable des relations UCS <-> variables (a partir des modeles)."""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from sklearn.inspection import permutation_importance
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.model_selection import train_test_split

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    # Permet d'importer src.* quand le script est lance directement.
    sys.path.insert(0, str(ROOT_DIR))

from src.schema import clean_dataframe, standardize_required_columns

TARGET_UCS = "UCS28d (kPa)"

PDP_FEATURES = [
    "P20 (\u00b5m)",
    "P80 (\u00b5m)",
    "Cw_f",
    "muscovite_total (%)",
    "muscovite_added (%)",
    "muscovite_ratio",
    "Ad %",
    "E/C",
]

INTERACTIONS = [
    ("P20 (\u00b5m)", "P80 (\u00b5m)"),
    ("Cw_f", "E/C"),
]


def _load_json(path: Path) -> dict:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _detect_sheet(path: Path) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if not wb.sheetnames:
            raise ValueError("Aucune feuille detectee.")
        return wb.sheetnames[0]
    finally:
        wb.close()


def _select_tailings(tailings_arg: str, df: pd.DataFrame | None, dataset_path: str | None) -> str:
    if tailings_arg and tailings_arg.upper() in {"L01", "WW"}:
        return tailings_arg.upper()
    if df is not None and "Tailings" in df.columns:
        uniques = sorted({str(v) for v in df["Tailings"].dropna().unique()})
        if len(uniques) == 1:
            return uniques[0].upper()
    if dataset_path:
        name = dataset_path.lower()
        if "l01" in name:
            return "L01"
        if "ww" in name:
            return "WW"
    return "L01"


def _load_model(models_dir: Path, tailings: str):
    if tailings == "WW":
        path = models_dir / "ww_ucs.joblib"
    else:
        path = models_dir / "l01_ucs.joblib"
    if not path.exists():
        path = models_dir / "ucs.joblib"
    if not path.exists():
        raise FileNotFoundError(f"Modele UCS introuvable: {path}")
    import joblib
    return joblib.load(path)


def _unwrap_model(model):
    try:
        from sklearn.compose import TransformedTargetRegressor
    except Exception:
        TransformedTargetRegressor = None
    if TransformedTargetRegressor and isinstance(model, TransformedTargetRegressor):
        return model.regressor_ if hasattr(model, "regressor_") else model.regressor
    return model


def _get_transformed_feature_names(model) -> list[str] | None:
    model = _unwrap_model(model)
    try:
        from sklearn.pipeline import Pipeline
    except Exception:
        Pipeline = None
    if Pipeline and isinstance(model, Pipeline):
        preprocess = model.named_steps.get("preprocess")
        if preprocess is not None and hasattr(preprocess, "get_feature_names_out"):
            return list(preprocess.get_feature_names_out())
    return None


def _extract_features(metadata: dict, tailings: str) -> tuple[list[str], list[str]]:
    features = metadata.get("features", {})
    section = features.get(tailings, {}) if tailings in features else features
    ucs_section = section.get("ucs", {}) if isinstance(section, dict) else {}
    cat_cols = ucs_section.get("categorical", []) or []
    num_cols = ucs_section.get("numeric", []) or []
    return list(cat_cols), list(num_cols)


def _compute_stats(df: pd.DataFrame, columns: list[str]) -> dict:
    stats = {}
    for col in columns:
        if col not in df.columns:
            continue
        series = pd.to_numeric(df[col], errors="coerce")
        if series.notna().sum() == 0:
            continue
        stats[col] = {
            "min": float(series.min()),
            "max": float(series.max()),
            "mean": float(series.mean()),
            "std": float(series.std()),
            "p05": float(series.quantile(0.05)),
            "p95": float(series.quantile(0.95)),
        }
    return stats


def _bounds_from_stats(stats: dict) -> tuple[float, float, str]:
    low = stats.get("p05")
    high = stats.get("p95")
    note = "p05-p95"
    if low is None or high is None or np.isnan(low) or np.isnan(high):
        low = stats.get("min")
        high = stats.get("max")
        note = "min-max"
    if low is None or high is None or np.isnan(low) or np.isnan(high):
        raise ValueError("Bornes invalides pour la grille PDP.")
    return float(low), float(high), note


def _grid_from_stats(stats: dict, n_points: int = 20) -> tuple[np.ndarray, str]:
    low, high, note = _bounds_from_stats(stats)
    if low == high:
        grid = np.array([low])
    else:
        grid = np.linspace(low, high, n_points)
    return grid, note


def _is_low_variation(stats: dict) -> bool:
    try:
        low, high, _ = _bounds_from_stats(stats)
    except ValueError:
        return False
    range_val = high - low
    if range_val <= 0:
        return True
    mean = stats.get("mean")
    if mean is None or np.isnan(mean):
        mean = 0.0
    scale = max(1.0, abs(float(mean)))
    return range_val <= 0.02 * scale


def _recompute_ratio(df: pd.DataFrame) -> None:
    if "muscovite_ratio" not in df.columns:
        return
    if "muscovite_added (%)" not in df.columns or "muscovite_total (%)" not in df.columns:
        return
    total = pd.to_numeric(df["muscovite_total (%)"], errors="coerce")
    added = pd.to_numeric(df["muscovite_added (%)"], errors="coerce")
    ratio = np.where(total > 0, added / total, 0.0)
    df["muscovite_ratio"] = ratio


def _safe_predict(model, X: pd.DataFrame) -> np.ndarray:
    return model.predict(X)


def _safe_token(text: str) -> str:
    """Nettoie un nom de feature pour creer un nom de fichier ASCII stable."""
    token = (
        text.replace(" ", "_")
        .replace("%", "pct")
        .replace("(", "")
        .replace(")", "")
        .replace("/", "_")
        .replace("\\", "_")
    )
    token = token.replace("\u00b5", "u").replace("\u03bc", "u")
    return token


def _save_csv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)


def _save_line_plot(x: np.ndarray, y: np.ndarray, title: str, path: Path) -> None:
    plt.figure(figsize=(6, 4))
    plt.plot(x, y, marker="o")
    plt.title(title)
    plt.xlabel("Valeur")
    plt.ylabel("UCS predit (kPa)")
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(path, dpi=150)
    plt.close()


def _save_bar_plot(labels: list[str], values: list[float], title: str, path: Path) -> None:
    plt.figure(figsize=(7, 4))
    y_pos = np.arange(len(labels))
    plt.barh(y_pos, values)
    plt.yticks(y_pos, labels)
    plt.xlabel("Importance")
    plt.title(title)
    plt.tight_layout()
    plt.savefig(path, dpi=150)
    plt.close()


def _save_heatmap(grid_x: np.ndarray, grid_y: np.ndarray, z: np.ndarray, title: str, path: Path) -> None:
    plt.figure(figsize=(6, 5))
    plt.imshow(z, origin="lower", aspect="auto",
               extent=[grid_x.min(), grid_x.max(), grid_y.min(), grid_y.max()])
    plt.colorbar(label="UCS predit (kPa)")
    plt.title(title)
    plt.xlabel("X")
    plt.ylabel("Y")
    plt.tight_layout()
    plt.savefig(path, dpi=150)
    plt.close()


def _infer_binders(df: pd.DataFrame | None, model) -> list[str]:
    if df is not None and "Binder" in df.columns:
        values = sorted({str(v) for v in df["Binder"].dropna().unique()})
        if values:
            return values
    return ["GUL", "20G80S", "Slag"]


def _compute_direction(grid: np.ndarray, preds: np.ndarray) -> str:
    if len(grid) < 2:
        return "variation insuffisante"
    corr = np.corrcoef(grid, preds)[0, 1]
    if np.isnan(corr):
        return "relation incertaine"
    if corr > 0.3:
        return "tendance positive"
    if corr < -0.3:
        return "tendance negative"
    return "tendance faible ou non lineaire"


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyse interpretable UCS.")
    parser.add_argument("--models-dir", required=True, help="Dossier des modeles finaux.")
    parser.add_argument("--dataset-xlsx", default="", help="Excel de reference (optionnel).")
    parser.add_argument("--tailings", default="auto", choices=["L01", "WW", "auto"], help="Tailings cible.")
    parser.add_argument("--out-dir", required=True, help="Dossier de sortie.")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--n-sample-pdp", type=int, default=200)
    args = parser.parse_args()

    np.random.seed(args.seed)

    models_dir = Path(args.models_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    metadata = _load_json(models_dir / "metadata.json")
    metrics = _load_json(models_dir / "metrics.json")

    df = None
    y = None
    dataset_path = args.dataset_xlsx or None
    if dataset_path:
        path = Path(dataset_path)
        sheet = _detect_sheet(path)
        raw = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
        raw = clean_dataframe(raw)
        raw = standardize_required_columns(raw)
        df = raw

    tailings = _select_tailings(args.tailings, df, dataset_path)
    cat_cols, num_cols = _extract_features(metadata, tailings)
    if not num_cols and not cat_cols:
        raise ValueError("Features UCS introuvables dans metadata.json.")

    if df is not None:
        if "Tailings" not in df.columns:
            df["Tailings"] = tailings
        if "Tailings" in df.columns:
            df = df[df["Tailings"].astype(str).str.upper() == tailings]
        missing = [c for c in num_cols + cat_cols if c not in df.columns]
        if missing:
            raise ValueError("Colonnes manquantes dans le dataset: " + ", ".join(missing))
        if TARGET_UCS in df.columns:
            y = df[TARGET_UCS].astype(float)
        X = df[num_cols + cat_cols].copy()
    else:
        X = None

    model = _load_model(models_dir, tailings)
    stats = _compute_stats(df, num_cols) if df is not None else {}

    report_lines: list[str] = []
    report_lines.append(f"# Rapport interpretabilite UCS ({tailings})")

    if metrics:
        ucs_metrics = metrics.get("UCS", {}).get("Tailings", {}).get(tailings, {})
        if not ucs_metrics:
            ucs_metrics = metrics.get("UCS", {}).get("overall", {})
        if ucs_metrics:
            report_lines.append(
                f"Performance (metrics sauvegardees): R^2={ucs_metrics.get('r2', 'n/a')}, RMSE={ucs_metrics.get('rmse', 'n/a')}"
            )

    if X is not None and y is not None:
        X_train, X_valid, y_train, y_valid = train_test_split(
            X, y, test_size=0.25, random_state=args.seed
        )
        try:
            perm = permutation_importance(
                model,
                X_valid,
                y_valid,
                n_repeats=10,
                random_state=args.seed,
                scoring="neg_root_mean_squared_error",
            )
            imp_df = pd.DataFrame(
                {
                    "feature": X_valid.columns,
                    "importance_mean": perm.importances_mean,
                    "importance_std": perm.importances_std,
                }
            ).sort_values("importance_mean", ascending=False)
            _save_csv(imp_df, out_dir / "importance_permutation.csv")
            _save_bar_plot(
                imp_df["feature"].tolist()[:15],
                imp_df["importance_mean"].tolist()[:15],
                "Importance (permutation)",
                out_dir / "importance_permutation.png",
            )
            report_lines.append("## Top 10 importances permutation")
            for _, row in imp_df.head(10).iterrows():
                report_lines.append(f"- {row['feature']}: {row['importance_mean']:.4f}")
        except Exception as exc:
            report_lines.append(f"Permutation importance indisponible: {exc}")

        # PDP 1D
        sample_size = min(len(X_train), args.n_sample_pdp)
        X_sample = X_train.sample(n=sample_size, random_state=args.seed).copy()
        report_lines.append("")
        report_lines.append("## Relations (PDP 1D approx)")
        for feature in PDP_FEATURES:
            if feature not in num_cols:
                continue
            stats_f = stats.get(feature, {})
            if not stats_f:
                continue
            grid, note = _grid_from_stats(stats_f, n_points=20)
            preds = []
            for val in grid:
                X_tmp = X_sample.copy()
                X_tmp[feature] = val
                if feature in {"muscovite_total (%)", "muscovite_added (%)"}:
                    _recompute_ratio(X_tmp)
                preds.append(np.mean(_safe_predict(model, X_tmp)))
            preds = np.asarray(preds)
            pdp_df = pd.DataFrame({"grid_value": grid, "pred_mean": preds})
            file_tag = _safe_token(feature)
            _save_csv(pdp_df, out_dir / f"pdp_{file_tag}.csv")
            _save_line_plot(grid, preds, f"PDP {feature}", out_dir / f"pdp_{file_tag}.png")

            direction = _compute_direction(grid, preds)
            if _is_low_variation(stats_f):
                report_lines.append(f"- {feature}: variation faible, relation incertaine.")
            else:
                extra = ", attention extrapolation" if note == "min-max" else ""
                report_lines.append(f"- {feature}: {direction} (grille {note}{extra}).")

        # Binder effect
        if "Binder" in cat_cols:
            binders = _infer_binders(df, model)
            if binders:
                binder_rows = []
                for binder in binders:
                    X_tmp = X_sample.copy()
                    X_tmp["Binder"] = binder
                    binder_rows.append({"Binder": binder, "pred_mean": float(np.mean(_safe_predict(model, X_tmp)))})
                binder_df = pd.DataFrame(binder_rows)
                _save_csv(binder_df, out_dir / "binder_effect.csv")
                _save_bar_plot(binder_df["Binder"].tolist(), binder_df["pred_mean"].tolist(), "Effet Binder", out_dir / "binder_effect.png")
                report_lines.append("")
                report_lines.append("## Effet Binder")
                for _, row in binder_df.iterrows():
                    report_lines.append(f"- {row['Binder']}: {row['pred_mean']:.2f} kPa")

        # Interactions 2D
        report_lines.append("")
        report_lines.append("## Interactions (PDP 2D)")
        for feat_x, feat_y in INTERACTIONS:
            if feat_x not in num_cols or feat_y not in num_cols:
                continue
            stats_x = stats.get(feat_x, {})
            stats_y = stats.get(feat_y, {})
            if not stats_x or not stats_y:
                continue
            grid_x, note_x = _grid_from_stats(stats_x, n_points=20)
            grid_y, note_y = _grid_from_stats(stats_y, n_points=20)
            z = np.zeros((len(grid_y), len(grid_x)))
            for iy, yv in enumerate(grid_y):
                for ix, xv in enumerate(grid_x):
                    X_tmp = X_sample.copy()
                    X_tmp[feat_x] = xv
                    X_tmp[feat_y] = yv
                    if feat_x in {"muscovite_total (%)", "muscovite_added (%)"} or feat_y in {"muscovite_total (%)", "muscovite_added (%)"}:
                        _recompute_ratio(X_tmp)
                    z[iy, ix] = float(np.mean(_safe_predict(model, X_tmp)))
            grid_df = pd.DataFrame(
                [
                    {"x": xv, "y": yv, "pred_mean": z[iy, ix]}
                    for iy, yv in enumerate(grid_y)
                    for ix, xv in enumerate(grid_x)
                ]
            )
            tag = f"{_safe_token(feat_x)}_x_{_safe_token(feat_y)}"
            _save_csv(grid_df, out_dir / f"pdp2d_{tag}.csv")
            _save_heatmap(grid_x, grid_y, z, f"Interaction {feat_x} vs {feat_y}", out_dir / f"pdp2d_{tag}.png")
            extrap = note_x == "min-max" or note_y == "min-max"
            extra = ", attention extrapolation" if extrap else ""
            report_lines.append(f"- Interaction {feat_x} vs {feat_y} (grille {note_x}/{note_y}{extra})")

        if not metrics:
            preds = _safe_predict(model, X_valid)
            r2 = r2_score(y_valid, preds)
            rmse = math.sqrt(mean_squared_error(y_valid, preds))
            report_lines.append("")
            report_lines.append(f"Performance holdout: R^2={r2:.3f}, RMSE={rmse:.2f}")
    else:
        report_lines.append("")
        report_lines.append("Aucune donnee brute fournie.")
        report_lines.append("- Permutation importance / PDP non calcules.")
        model_base = _unwrap_model(model)
        importances = getattr(model_base, "feature_importances_", None)
        names = _get_transformed_feature_names(model)
        if importances is not None:
            if names is None:
                names = [f"feature_{i}" for i in range(len(importances))]
            imp_df = pd.DataFrame(
                {
                    "feature": names,
                    "importance_mean": importances,
                }
            ).sort_values("importance_mean", ascending=False)
            _save_csv(imp_df, out_dir / "importance_model.csv")
            _save_bar_plot(
                imp_df["feature"].tolist()[:15],
                imp_df["importance_mean"].tolist()[:15],
                "Importance (modele, features transformees)",
                out_dir / "importance_model.png",
            )
            report_lines.append("- Importance modele exportee (features transformees).")

    report_lines.append("")
    report_lines.append("## Limitations")
    report_lines.append("- Associations apprises par le modele, pas une causalite.")
    report_lines.append("- Variation faible => relation moins fiable.")

    report_path = out_dir / "rapport.md"
    report_path.write_text("\n".join(report_lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
