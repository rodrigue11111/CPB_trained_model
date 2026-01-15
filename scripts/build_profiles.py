"""Build minimal data profiles for Streamlit Cloud.

This script reads local Excel files (data/) and produces a small JSON
with numeric stats + categorical values. It avoids shipping raw data
to the repository.
"""

from __future__ import annotations

from pathlib import Path
import json
import sys
import random

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.schema import clean_dataframe, standardize_required_columns


DATASETS = {
    "WW": {
        "path": "data/WW-Optimisation.xlsx",
        "sheet": "Feuil3 (3)",
        "tailings": "WW",
    },
    "L01_OLD": {
        "path": "data/L01-Optimisation.xlsx",
        "sheet": "Feuil1 (2)",
        "tailings": "L01",
    },
    "L01_NEW": {
        "path": "data/L01-dataset.xlsx",
        "sheet": None,
        "tailings": "L01",
    },
}

SEED = 42
SAMPLE_FRACTION = 0.5
MAX_VALUES_PER_COL = 5000


def _first_sheet(path: str | Path) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if not wb.sheetnames:
            raise ValueError("No sheets found.")
        return wb.sheetnames[0]
    finally:
        wb.close()


def _compute_stats(df: pd.DataFrame) -> dict:
    stats = {}
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            series = pd.to_numeric(df[col], errors="coerce")
            stats[col] = {
                "min": float(series.min(skipna=True)) if series.notna().any() else None,
                "max": float(series.max(skipna=True)) if series.notna().any() else None,
                "mean": float(series.mean(skipna=True)) if series.notna().any() else None,
                "std": float(series.std(skipna=True)) if series.notna().any() else None,
            }
    return stats


def _sample_values(series: pd.Series) -> list[float]:
    values = pd.to_numeric(series, errors="coerce").dropna().tolist()
    if not values:
        return []
    n = max(1, int(len(values) * SAMPLE_FRACTION))
    if n >= len(values):
        sampled = values
    else:
        sampled = random.sample(values, n)
    if len(sampled) > MAX_VALUES_PER_COL:
        sampled = random.sample(sampled, MAX_VALUES_PER_COL)
    return [float(v) for v in sampled]


def build_profiles() -> dict:
    profiles: dict = {}
    random.seed(SEED)
    for key, cfg in DATASETS.items():
        path = Path(cfg["path"])
        if not path.exists():
            print(f"[WARN] Missing file: {path}")
            continue

        sheet = cfg["sheet"]
        if sheet is None:
            sheet = _first_sheet(path)

        df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
        df["Tailings"] = cfg["tailings"]
        df = clean_dataframe(df)
        df = standardize_required_columns(df)

        numeric_stats = _compute_stats(df)
        sample_values: dict[str, list[float]] = {}
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                sample_values[col] = _sample_values(df[col])
        categorical_values = {}
        for col in df.columns:
            if col in {"Binder", "Tailings"}:
                values = sorted({str(v) for v in df[col].dropna().unique()})
                categorical_values[col] = values

        profiles[key] = {
            "numeric_stats": numeric_stats,
            "categorical_values": categorical_values,
            "sample_values": sample_values,
        }
        print(f"[OK] Built profile for {key} with {len(numeric_stats)} numeric cols.")

    return profiles


def main() -> int:
    profiles = build_profiles()
    if not profiles:
        print("No profiles generated.")
        return 1

    out_path = Path("app/assets/data_profiles.json")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(profiles, ensure_ascii=True, indent=2), encoding="utf-8"
    )
    print(f"Wrote: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
