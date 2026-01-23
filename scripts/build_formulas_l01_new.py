# -*- coding: utf-8 -*-
"""Construit des formules interpretable UCS pour L01 NEW (global + regles)."""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd

from sklearn.compose import ColumnTransformer
from sklearn.linear_model import ElasticNet, LinearRegression, RidgeCV
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.model_selection import GridSearchCV, KFold, train_test_split
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder, SplineTransformer, StandardScaler
from sklearn.tree import DecisionTreeRegressor, export_text
from sklearn.impute import SimpleImputer

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    # Permet d'importer src.* quand le script est lance directement.
    sys.path.insert(0, str(ROOT_DIR))

from src.schema import clean_dataframe, standardize_required_columns
from src.features import infer_feature_columns

TARGET_UCS = "UCS28d (kPa)"


def _load_json(path: Path) -> dict:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _detect_sheet(path: Path) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if not wb.sheetnames:
            raise ValueError("Aucune feuille detectee.")
        return wb.sheetnames[0]
    finally:
        wb.close()


def _parse_floats(text: str) -> list[float]:
    values = []
    for item in (text or "").split(","):
        item = item.strip()
        if not item:
            continue
        values.append(float(item))
    return values


def _ensure_ratio(df: pd.DataFrame) -> pd.DataFrame:
    if "muscovite_ratio" in df.columns:
        return df
    if "muscovite_added (%)" in df.columns and "muscovite_total (%)" in df.columns:
        total = pd.to_numeric(df["muscovite_total (%)"], errors="coerce")
        added = pd.to_numeric(df["muscovite_added (%)"], errors="coerce")
        df = df.copy()
        df["muscovite_ratio"] = np.where(total > 0, added / total, 0.0)
    return df


def _select_features(
    df: pd.DataFrame,
    metadata: dict | None,
) -> tuple[list[str], list[str]]:
    if metadata:
        section = metadata.get("features", {}).get("L01", {})
        ucs_section = section.get("ucs", {}) if isinstance(section, dict) else {}
        cat_cols = list(ucs_section.get("categorical", []) or [])
        num_cols = list(ucs_section.get("numeric", []) or [])
        if cat_cols or num_cols:
            return cat_cols, num_cols
    # Fallback: selection dynamique
    cat_cols, num_cols = infer_feature_columns(df, target_cols=[TARGET_UCS])
    if "Binder" in df.columns and "Binder" not in cat_cols:
        cat_cols.append("Binder")
    return cat_cols, num_cols


def _prepare_xy(df: pd.DataFrame, cat_cols: list[str], num_cols: list[str]):
    missing = [c for c in cat_cols + num_cols if c not in df.columns]
    if missing:
        raise ValueError("Colonnes manquantes: " + ", ".join(missing))
    X = df[cat_cols + num_cols].copy()
    y = df[TARGET_UCS].astype(float).copy()
    return X, y


def _safe_token(text: str) -> str:
    token = (
        text.replace(" ", "_")
        .replace("%", "pct")
        .replace("(", "")
        .replace(")", "")
        .replace("/", "_")
        .replace("\\", "_")
    )
    token = token.replace("\u00b5", "u").replace("\u03bc", "u")
    return token


def _infer_source(term: str, numeric_cols: Iterable[str], categorical_cols: Iterable[str]) -> str:
    term_lower = "".join([c for c in term.lower() if c.isalnum()])
    for col in numeric_cols:
        key = "".join([c for c in col.lower() if c.isalnum()])
        if key and key in term_lower:
            return col
    for col in categorical_cols:
        key = "".join([c for c in col.lower() if c.isalnum()])
        if key and key in term_lower:
            return col
    if "binder" in term.lower():
        return "Binder"
    return "unknown"


def _extract_coefficients(
    model: Pipeline,
    numeric_cols: list[str],
    categorical_cols: list[str],
) -> pd.DataFrame:
    preprocessor = model.named_steps["preprocess"]
    estimator = model.named_steps["model"]
    names = []
    if hasattr(preprocessor, "get_feature_names_out"):
        names = list(preprocessor.get_feature_names_out())
    coefs = estimator.coef_
    rows = []
    for name, coef in zip(names, coefs):
        clean = name.replace("num__", "").replace("cat__", "")
        rows.append(
            {
                "term_name": clean,
                "coefficient": float(coef),
                "feature_source": _infer_source(clean, numeric_cols, categorical_cols),
                "knot_info": "",
            }
        )
    rows.append(
        {
            "term_name": "intercept",
            "coefficient": float(estimator.intercept_),
            "feature_source": "intercept",
            "knot_info": "",
        }
    )
    df = pd.DataFrame(rows)
    df["abs_coef"] = df["coefficient"].abs()
    df = df.sort_values("abs_coef", ascending=False).drop(columns=["abs_coef"])
    return df


def _extract_classic_coefficients(
    model: Pipeline,
    numeric_cols: list[str],
    categorical_cols: list[str],
) -> tuple[pd.DataFrame, float]:
    """Retourne les coefficients en unites originales + intercept ajuste."""
    preprocessor = model.named_steps["preprocess"]
    estimator = model.named_steps["model"]
    names = []
    if hasattr(preprocessor, "get_feature_names_out"):
        names = list(preprocessor.get_feature_names_out())
    coef_std = estimator.coef_.ravel()
    coef_map = dict(zip(names, coef_std))

    intercept = float(estimator.intercept_)
    intercept_adjust = 0.0

    coef_rows = []
    if "num" in preprocessor.named_transformers_:
        num_pipe = preprocessor.named_transformers_["num"]
        if isinstance(num_pipe, Pipeline):
            scaler = num_pipe.named_steps.get("scaler")
            if scaler is not None:
                means = scaler.mean_
                scales = scaler.scale_
                for idx, col in enumerate(numeric_cols):
                    key = f"num__{col}"
                    coef_val = float(coef_map.get(key, 0.0))
                    scale = float(scales[idx]) if scales is not None else 1.0
                    mean = float(means[idx]) if means is not None else 0.0
                    if scale == 0:
                        coef_orig = 0.0
                    else:
                        coef_orig = coef_val / scale
                        intercept_adjust += coef_val * mean / scale
                    coef_rows.append(
                        {
                            "term_name": col,
                            "coefficient": float(coef_orig),
                            "feature_source": col,
                            "term_type": "numeric",
                        }
                    )

    # Coefficients categiels (pas de scaling)
    for name in names:
        if not name.startswith("cat__"):
            continue
        coef_val = float(coef_map.get(name, 0.0))
        clean = name.replace("cat__", "")
        coef_rows.append(
            {
                "term_name": clean,
                "coefficient": coef_val,
                "feature_source": "Binder",
                "term_type": "categorical",
            }
        )

    intercept_original = intercept - intercept_adjust
    coef_rows.append(
        {
            "term_name": "intercept",
            "coefficient": float(intercept_original),
            "feature_source": "intercept",
            "term_type": "intercept",
        }
    )

    df = pd.DataFrame(coef_rows)
    df["abs_coef"] = df["coefficient"].abs()
    df = df.sort_values("abs_coef", ascending=False).drop(columns=["abs_coef"])
    return df, float(intercept_original)


def _format_equation_lines(
    title: str,
    intercept: float,
    coef_df: pd.DataFrame,
    metrics: dict,
    features: list[str],
    note: str,
) -> list[str]:
    lines = [
        f"# {title}",
        "",
        "Forme g\u00e9n\u00e9rale :",
        "UCS = intercept + somme(coef * variable)",
        "",
        f"Intercept = {intercept:.4f}",
        "",
        "Termes :",
    ]
    for _, row in coef_df.iterrows():
        if row["term_name"] == "intercept":
            continue
        term = row["term_name"]
        coef = row["coefficient"]
        if term.startswith("Binder_"):
            label = term.replace("Binder_", "Binder=")
        else:
            label = term
        lines.append(f"- {coef:+.4f} * {label}")
    lines.append("")
    lines.append("Interpr\u00e9tation rapide :")
    lines.append("- Un coefficient positif augmente UCS quand la variable augmente.")
    lines.append("- Un coefficient n\u00e9gatif diminue UCS quand la variable augmente.")
    lines.append("")
    lines.append(
        f"Performance test : R^2 = {metrics.get('r2'):.3f}, RMSE = {metrics.get('rmse'):.2f} kPa"
    )
    lines.append("")
    lines.append("Features retenues :")
    for feat in features:
        lines.append(f"- {feat}")
    lines.append("")
    lines.append(
        "Avertissement : \u00e9quation valable dans la plage des donn\u00e9es L01 NEW."
    )
    lines.append(note)
    return lines


def _write_rules(path: Path, title: str, rules: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = [
        title,
        "",
        "Comment lire les regles :",
        "- Chaque ligne est une condition sur une variable.",
        "- La derniere ligne indique la prediction UCS (kPa).",
        "",
        rules,
        "",
    ]
    path.write_text("\n".join(text), encoding="utf-8")


def _safe_import_matplotlib():
    try:
        import matplotlib.pyplot as plt
    except Exception:
        return None
    return plt


def _plot_scatter(plt, x, y, title, path: Path):
    fig, ax = plt.subplots(figsize=(5, 4))
    ax.scatter(x, y, alpha=0.7)
    ax.plot([x.min(), x.max()], [x.min(), x.max()], "--", color="#999999")
    ax.set_xlabel("UCS reel")
    ax.set_ylabel("UCS predit")
    ax.set_title(title)
    fig.tight_layout()
    fig.savefig(path, dpi=150)
    plt.close(fig)


def _plot_residuals(plt, preds, residuals, title, path: Path):
    fig, ax = plt.subplots(figsize=(5, 4))
    ax.scatter(preds, residuals, alpha=0.7)
    ax.axhline(0, color="#999999", linestyle="--")
    ax.set_xlabel("UCS predit")
    ax.set_ylabel("Residus (reel - predit)")
    ax.set_title(title)
    fig.tight_layout()
    fig.savefig(path, dpi=150)
    plt.close(fig)


def main() -> None:
    parser = argparse.ArgumentParser(description="Formules UCS L01 NEW.")
    parser.add_argument("--dataset-xlsx", required=True)
    parser.add_argument("--models-dir", default="", help="Optionnel: modele L01 NEW pour la fidelite.")
    parser.add_argument("--out-dir", default="outputs/formulas/L01_new")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--test-size", type=float, default=0.25)
    parser.add_argument("--max-depth", type=int, default=4)
    parser.add_argument("--splines-knots", type=int, default=6)
    parser.add_argument("--splines-degree", type=int, default=3)
    parser.add_argument("--alpha-grid", default="1e-4,1e-3,1e-2,1e-1,1")
    parser.add_argument("--l1-ratio-grid", default="0.05,0.1,0.3,0.5,0.7,0.9")
    parser.add_argument("--max-nonzero", type=int, default=12)
    args = parser.parse_args()

    np.random.seed(args.seed)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    dataset_path = Path(args.dataset_xlsx)
    sheet = _detect_sheet(dataset_path)
    raw = pd.read_excel(dataset_path, sheet_name=sheet, engine="openpyxl")
    raw = clean_dataframe(raw)
    raw = standardize_required_columns(raw)
    raw = _ensure_ratio(raw)

    if "Tailings" in raw.columns:
        raw = raw[raw["Tailings"].astype(str).str.upper() == "L01"]
    raw = raw[raw[TARGET_UCS].notna()].copy()

    metadata = _load_json(Path(args.models_dir) / "metadata.json") if args.models_dir else {}
    cat_cols, num_cols = _select_features(raw, metadata)

    X, y = _prepare_xy(raw, cat_cols, num_cols)
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=args.test_size, random_state=args.seed
    )

    # ------------------------------------------------------
    # CLASSIC LINEAR (sans splines) - equation interpretable
    # ------------------------------------------------------
    numeric_pipe = Pipeline(
        steps=[
            ("imputer", SimpleImputer(strategy="median")),
            ("scaler", StandardScaler()),
        ]
    )
    cat_pipe = Pipeline(
        steps=[
            ("imputer", SimpleImputer(strategy="most_frequent")),
            ("onehot", OneHotEncoder(handle_unknown="ignore")),
        ]
    )
    transformers = []
    if num_cols:
        transformers.append(("num", numeric_pipe, num_cols))
    if cat_cols:
        transformers.append(("cat", cat_pipe, cat_cols))
    classic_preprocess = ColumnTransformer(
        transformers=transformers,
        remainder="drop",
    )

    classic_models = {}
    classic_metrics = {}
    classic_preds = {}

    # A) LinearRegression baseline
    lr_model = Pipeline(
        steps=[
            ("preprocess", classic_preprocess),
            ("model", LinearRegression()),
        ]
    )
    lr_model.fit(X_train, y_train)
    pred_lr = lr_model.predict(X_test)
    classic_models["linear"] = lr_model
    classic_preds["linear"] = pred_lr
    classic_metrics["linear"] = {
        "r2": float(r2_score(y_test, pred_lr)),
        "rmse": float(math.sqrt(mean_squared_error(y_test, pred_lr))),
    }

    # B) RidgeCV (stabilite)
    ridge_alphas = np.logspace(-3, 3, 13)
    ridge_model = Pipeline(
        steps=[
            ("preprocess", classic_preprocess),
            ("model", RidgeCV(alphas=ridge_alphas)),
        ]
    )
    ridge_model.fit(X_train, y_train)
    pred_ridge = ridge_model.predict(X_test)
    classic_models["ridge"] = ridge_model
    classic_preds["ridge"] = pred_ridge
    classic_metrics["ridge"] = {
        "r2": float(r2_score(y_test, pred_ridge)),
        "rmse": float(math.sqrt(mean_squared_error(y_test, pred_ridge))),
        "alpha": float(ridge_model.named_steps["model"].alpha_),
    }

    # C) ElasticNet (formule plus courte)
    enet_model = Pipeline(
        steps=[
            ("preprocess", classic_preprocess),
            ("model", ElasticNet(max_iter=10000, random_state=args.seed)),
        ]
    )
    alpha_grid = _parse_floats(args.alpha_grid)
    l1_grid = _parse_floats(args.l1_ratio_grid)
    n_splits = min(5, len(X_train))
    if n_splits < 2:
        n_splits = 2
    cv = KFold(n_splits=n_splits, shuffle=True, random_state=args.seed)
    enet_grid = GridSearchCV(
        enet_model,
        param_grid={
            "model__alpha": alpha_grid,
            "model__l1_ratio": l1_grid,
        },
        scoring="neg_root_mean_squared_error",
        cv=cv,
    )
    enet_grid.fit(X_train, y_train)
    enet_best = enet_grid.best_estimator_
    pred_enet = enet_best.predict(X_test)
    classic_models["elasticnet"] = enet_best
    classic_preds["elasticnet"] = pred_enet
    classic_metrics["elasticnet"] = {
        "r2": float(r2_score(y_test, pred_enet)),
        "rmse": float(math.sqrt(mean_squared_error(y_test, pred_enet))),
        "best_params": enet_grid.best_params_,
    }

    # Choix du meilleur modele (RMSE minimal)
    best_key = min(classic_metrics.keys(), key=lambda k: classic_metrics[k]["rmse"])
    best_classic = classic_models[best_key]

    # Export coefficients en unites originales
    coef_df, intercept_original = _extract_classic_coefficients(
        best_classic, num_cols, cat_cols
    )
    coef_df.to_csv(out_dir / "classic_linear_coefficients.csv", index=False)

    best_metrics = classic_metrics[best_key]
    classic_summary = {
        "best_model": best_key,
        "metrics": classic_metrics,
    }

    # Equation lisible (unites originales)
    equation_lines = _format_equation_lines(
        "Equation classique (L01 NEW)",
        intercept_original,
        coef_df,
        best_metrics,
        num_cols + cat_cols,
        "",
    )
    (out_dir / "classic_linear_equation.md").write_text(
        "\n".join(equation_lines), encoding="utf-8"
    )

    import joblib

    joblib.dump(best_classic, out_dir / "classic_linear.joblib")
    # On ecrit les metriques plus tard (apres la fidelite si disponible).

    # Option equation courte (max nonzero)
    max_nonzero = getattr(args, "max_nonzero", 12)
    if best_key == "elasticnet":
        enet_pipe = classic_models["elasticnet"]
        coef_df_enet, intercept_enet = _extract_classic_coefficients(
            enet_pipe, num_cols, cat_cols
        )
        nonzero = (coef_df_enet["term_type"] != "intercept") & (coef_df_enet["coefficient"].abs() > 1e-8)
        if nonzero.sum() > max_nonzero:
            short_model = None
            for alpha in sorted(alpha_grid, reverse=True):
                for l1_ratio in sorted(l1_grid, reverse=True):
                    trial = Pipeline(
                        steps=[
                            ("preprocess", classic_preprocess),
                            ("model", ElasticNet(max_iter=10000, random_state=args.seed, alpha=alpha, l1_ratio=l1_ratio)),
                        ]
                    )
                    trial.fit(X_train, y_train)
                    coef_df_trial, intercept_trial = _extract_classic_coefficients(
                        trial, num_cols, cat_cols
                    )
                    nz = (coef_df_trial["term_type"] != "intercept") & (coef_df_trial["coefficient"].abs() > 1e-8)
                    if nz.sum() <= max_nonzero:
                        short_model = trial
                        coef_df_enet = coef_df_trial
                        intercept_enet = intercept_trial
                        break
                if short_model is not None:
                    break
            if short_model is not None:
                pred_short = short_model.predict(X_test)
                short_metrics = {
                    "r2": float(r2_score(y_test, pred_short)),
                    "rmse": float(math.sqrt(mean_squared_error(y_test, pred_short))),
                    "alpha": float(short_model.named_steps["model"].alpha),
                    "l1_ratio": float(short_model.named_steps["model"].l1_ratio),
                }
                short_lines = _format_equation_lines(
                    "Equation courte (ElasticNet)",
                    intercept_enet,
                    coef_df_enet,
                    short_metrics,
                    num_cols + cat_cols,
                    f"Nombre de termes (hors intercept) <= {max_nonzero}",
                )
                (out_dir / "classic_linear_equation_short.md").write_text(
                    "\n".join(short_lines), encoding="utf-8"
                )

    # A) Modele global spline + ElasticNet
    numeric_transformer = SplineTransformer(
        n_knots=args.splines_knots,
        degree=args.splines_degree,
        include_bias=False,
    )
    categorical_transformer = OneHotEncoder(handle_unknown="ignore")
    preprocessor = ColumnTransformer(
        transformers=[
            ("num", numeric_transformer, num_cols),
            ("cat", categorical_transformer, cat_cols),
        ],
        remainder="drop",
    )

    pipeline = Pipeline(
        steps=[
            ("preprocess", preprocessor),
            ("model", ElasticNet(max_iter=5000, random_state=args.seed)),
        ]
    )

    alpha_grid = _parse_floats(args.alpha_grid)
    l1_grid = _parse_floats(args.l1_ratio_grid)
    n_splits = min(5, len(X_train))
    if n_splits < 2:
        n_splits = 2
    cv = KFold(n_splits=n_splits, shuffle=True, random_state=args.seed)
    grid = GridSearchCV(
        pipeline,
        param_grid={
            "model__alpha": alpha_grid,
            "model__l1_ratio": l1_grid,
        },
        scoring="neg_root_mean_squared_error",
        cv=cv,
    )
    grid.fit(X_train, y_train)
    best_model = grid.best_estimator_

    pred_global = best_model.predict(X_test)
    r2_global = r2_score(y_test, pred_global)
    rmse_global = math.sqrt(mean_squared_error(y_test, pred_global))

    global_metrics = {
        "r2": float(r2_global),
        "rmse": float(rmse_global),
        "best_params": grid.best_params_,
    }

    import joblib

    joblib.dump(best_model, out_dir / "global_spline_enet.joblib")
    (out_dir / "global_spline_enet_metrics.json").write_text(
        json.dumps(global_metrics, indent=2), encoding="utf-8"
    )

    coef_df = _extract_coefficients(best_model, num_cols, cat_cols)
    coef_df.to_csv(out_dir / "global_spline_enet_coefficients.csv", index=False)

    top_terms = coef_df.head(30)
    formula_lines = [
        "# Formule globale (Spline + ElasticNet)",
        "",
        "Forme generale :",
        "UCS = b0 + somme(coeff * bases_splines) + effets du liant (Binder).",
        "",
        "Meilleurs parametres :",
        f"- alpha = {grid.best_params_.get('model__alpha')}",
        f"- l1_ratio = {grid.best_params_.get('model__l1_ratio')}",
        "",
        f"Performance test : R^2 = {r2_global:.3f}, RMSE = {rmse_global:.2f} kPa",
        "",
        "Top 30 termes (coefficients les plus influents) :",
    ]
    for _, row in top_terms.iterrows():
        formula_lines.append(
            f"- {row['term_name']}: {row['coefficient']:.4f}"
        )
    formula_lines.append("")
    formula_lines.append(
        "Limites: valable dans la plage des donnees L01 NEW. "
        "Hors plage, la formule extrapole."
    )
    (out_dir / "global_formula.md").write_text("\n".join(formula_lines), encoding="utf-8")

    # B) Regles par arbre de decision
    rules_dir = out_dir
    tree_overall = DecisionTreeRegressor(max_depth=args.max_depth, random_state=args.seed)

    # Pour l'arbre global, on encode Binder si present.
    if cat_cols:
        tree_preprocess = ColumnTransformer(
            transformers=[
                ("num", "passthrough", num_cols),
                ("cat", OneHotEncoder(handle_unknown="ignore"), cat_cols),
            ]
        )
        X_train_tree = tree_preprocess.fit_transform(X_train)
        X_test_tree = tree_preprocess.transform(X_test)
        tree_overall.fit(X_train_tree, y_train)
        feature_names = []
        if hasattr(tree_preprocess, "get_feature_names_out"):
            feature_names = list(tree_preprocess.get_feature_names_out())
        rules_text = export_text(tree_overall, feature_names=feature_names)
        pred_tree = tree_overall.predict(X_test_tree)
    else:
        tree_overall.fit(X_train[num_cols], y_train)
        rules_text = export_text(tree_overall, feature_names=num_cols)
        pred_tree = tree_overall.predict(X_test[num_cols])

    _write_rules(rules_dir / "rules_tree_overall.txt", "Regles arbre global", rules_text)

    # Arbres par Binder (si applicable)
    binder_rules = {}
    binder_models = {}
    pred_rules = pd.Series(index=y_test.index, dtype=float)
    if "Binder" in X_train.columns:
        for binder in sorted(X_train["Binder"].dropna().unique()):
            binder_mask = X_train["Binder"] == binder
            tree = DecisionTreeRegressor(max_depth=args.max_depth, random_state=args.seed)
            tree.fit(X_train.loc[binder_mask, num_cols], y_train.loc[binder_mask])
            binder_models[binder] = tree
            binder_rules[binder] = export_text(tree, feature_names=num_cols)

        for binder in sorted(X_test["Binder"].dropna().unique()):
            binder_mask = X_test["Binder"] == binder
            tree = binder_models.get(binder)
            if tree is not None:
                pred_rules.loc[binder_mask] = tree.predict(X_test.loc[binder_mask, num_cols])

        for binder, rules in binder_rules.items():
            token = _safe_token(str(binder))
            _write_rules(
                rules_dir / f"rules_tree_by_binder_{token}.txt",
                f"Regles arbre - Binder {binder}",
                rules,
            )
    else:
        pred_rules = pd.Series(pred_tree, index=y_test.index)

    if pred_rules.isna().any():
        pred_rules = pred_rules.fillna(pd.Series(pred_tree, index=y_test.index))

    r2_tree = r2_score(y_test, pred_rules)
    rmse_tree = math.sqrt(mean_squared_error(y_test, pred_rules))

    tree_metrics = {
        "overall_tree": {
            "r2": float(r2_score(y_test, pred_tree)),
            "rmse": float(math.sqrt(mean_squared_error(y_test, pred_tree))),
        },
        "by_binder": {"r2": float(r2_tree), "rmse": float(rmse_tree)},
    }
    (out_dir / "tree_metrics.json").write_text(
        json.dumps(tree_metrics, indent=2), encoding="utf-8"
    )

    # Fidelite vs best model (si present)
    fidelity = {}
    comparison_rows = pd.DataFrame(
        {
            "y_true": y_test,
            "global_pred": pred_global,
            "rules_pred": pred_rules,
        }
    )
    if args.models_dir:
        model_path = Path(args.models_dir) / "l01_ucs.joblib"
        if model_path.exists():
            best_model = joblib.load(model_path)
            best_pred = best_model.predict(X_test)
            comparison_rows["best_pred"] = best_pred
            fidelity = {
                "global_vs_best": {
                    "r2": float(r2_score(best_pred, pred_global)),
                    "rmse": float(math.sqrt(mean_squared_error(best_pred, pred_global))),
                },
                "rules_vs_best": {
                    "r2": float(r2_score(best_pred, pred_rules)),
                    "rmse": float(math.sqrt(mean_squared_error(best_pred, pred_rules))),
                },
            }
            for key, preds in classic_preds.items():
                fidelity[f"classic_{key}_vs_best"] = {
                    "r2": float(r2_score(best_pred, preds)),
                    "rmse": float(math.sqrt(mean_squared_error(best_pred, preds))),
                }
            (out_dir / "fidelity_metrics.json").write_text(
                json.dumps(fidelity, indent=2), encoding="utf-8"
            )

    classic_summary["fidelity"] = fidelity
    (out_dir / "classic_linear_metrics.json").write_text(
        json.dumps(classic_summary, indent=2), encoding="utf-8"
    )

    comparison_rows.to_csv(out_dir / "comparaison_table.csv", index=False)

    # Figures
    fig_dir = out_dir / "figures"
    fig_dir.mkdir(parents=True, exist_ok=True)
    plt = _safe_import_matplotlib()
    if plt is None:
        warn = "matplotlib absent: figures non generees."
        (fig_dir / "README.txt").write_text(warn, encoding="utf-8")
    else:
        _plot_scatter(plt, y_test, pred_global, "Global (Spline + ENet)", fig_dir / "scatter_global.png")
        _plot_scatter(plt, y_test, pred_rules, "Regles (arbre)", fig_dir / "scatter_rules.png")
        _plot_residuals(plt, pred_global, y_test - pred_global, "Residus global", fig_dir / "residuals_global.png")
        _plot_residuals(plt, pred_rules, y_test - pred_rules, "Residus regles", fig_dir / "residuals_rules.png")

        if "best_pred" in comparison_rows.columns:
            _plot_scatter(plt, y_test, comparison_rows["best_pred"], "Modele best", fig_dir / "scatter_best.png")
            _plot_residuals(
                plt,
                comparison_rows["best_pred"],
                y_test - comparison_rows["best_pred"],
                "Residus best",
                fig_dir / "residuals_best.png",
            )


if __name__ == "__main__":
    main()
