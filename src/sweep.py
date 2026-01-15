"""Sweep multi-datasets et generation de tableaux de synthese.

Ce module lance de nombreux runs en appelant la CLI, puis consolide les
metriques dans des fichiers CSV (par run et par config).
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import statistics
import subprocess
import sys
import time
from itertools import product
from pathlib import Path

from openpyxl import load_workbook

CONFIG_FIELDS = [
    "dataset_id",
    "search_mode",
    "ww_ucs_model",
    "ww_ucs_transform",
    "ww_ucs_outliers",
    "ww_tune",
    "l01_ucs_model",
    "l01_ucs_transform",
    "l01_ucs_outliers",
    "l01_tune",
    "ww_slump_model",
    "ww_slump_tune",
    "l01_slump_model",
    "l01_slump_tune",
]

RUN_FIELDS = [
    "dataset_id",
    "run_id",
    "seed",
    "search_mode",
    "ww_ucs_model",
    "ww_ucs_transform",
    "ww_ucs_outliers",
    "ww_tune",
    "l01_ucs_model",
    "l01_ucs_transform",
    "l01_ucs_outliers",
    "l01_tune",
    "ww_slump_model",
    "ww_slump_tune",
    "l01_slump_model",
    "l01_slump_tune",
    "ucs_r2_l01",
    "ucs_rmse_l01",
    "ucs_r2_ww",
    "ucs_rmse_ww",
    "slump_r2_overall",
    "slump_rmse_overall",
    "pass_rate_pct",
    "run_score",
]

AGG_METRICS = [
    "run_score",
    "ucs_r2_l01",
    "ucs_rmse_l01",
    "ucs_r2_ww",
    "ucs_rmse_ww",
    "pass_rate_pct",
]


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Parse les arguments CLI pour piloter le sweep."""
    parser = argparse.ArgumentParser(
        description="Lance un sweep multi-datasets et compare les configurations."
    )
    parser.add_argument(
        "--out-dir",
        default="outputs/sweeps",
        help="Dossier de sortie principal.",
    )
    parser.add_argument(
        "--time-budget-min",
        type=float,
        default=300,
        help="Budget de temps en minutes.",
    )
    parser.add_argument(
        "--max-runs",
        type=int,
        default=0,
        help="Limite le nombre de runs (0 = pas de limite).",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Reprend les runs avec metrics.json existant.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Affiche le nombre de runs prevus sans executer.",
    )
    parser.add_argument(
        "--jobs",
        type=int,
        default=1,
        help="Nombre de jobs (execution sequentielle si >1).",
    )
    parser.add_argument(
        "--seeds",
        default="42,123",
        help="Liste de seeds separees par des virgules.",
    )
    parser.add_argument(
        "--search-modes",
        default="bootstrap",
        help="Liste des modes de recherche (uniform/bootstrap).",
    )
    parser.add_argument(
        "--n-samples",
        type=int,
        default=30000,
        help="Nombre de candidats par groupe Tailings/Binder.",
    )
    parser.add_argument(
        "--slump-min",
        type=float,
        default=70,
        help="Seuil minimum de Slump (mm) predit.",
    )
    parser.add_argument(
        "--ucs-min",
        type=float,
        default=900,
        help="Seuil minimum de UCS28d (kPa) predit.",
    )
    parser.add_argument(
        "--datasets",
        default="",
        help="Chemin d'un JSON de scenarios (optionnel).",
    )
    parser.add_argument(
        "--only-dataset",
        default="",
        help="Lance uniquement un dataset_id (ex: new).",
    )
    parser.add_argument(
        "--ww-ucs-models",
        default="gbr,hgb",
        help="Modeles UCS WW (liste).",
    )
    parser.add_argument(
        "--ww-ucs-transforms",
        default="none",
        help="Transformations UCS WW (liste).",
    )
    parser.add_argument(
        "--ww-ucs-outliers",
        default="none",
        help="Filtre outliers UCS WW (liste).",
    )
    parser.add_argument(
        "--ww-tune-options",
        default="false",
        help="Options tuning WW (liste true/false).",
    )
    parser.add_argument(
        "--l01-ucs-models",
        default="rf,et,hgb",
        help="Modeles UCS L01 (liste).",
    )
    parser.add_argument(
        "--l01-ucs-transforms",
        default="none,log",
        help="Transformations UCS L01 (liste).",
    )
    parser.add_argument(
        "--l01-ucs-outliers",
        default="none,iqr",
        help="Filtre outliers UCS L01 (liste).",
    )
    parser.add_argument(
        "--l01-tune-options",
        default="true",
        help="Options tuning L01 (liste true/false).",
    )
    parser.add_argument(
        "--ww-slump-models",
        default="gbr",
        help="Modeles Slump WW (liste).",
    )
    parser.add_argument(
        "--ww-slump-tune-options",
        default="false",
        help="Options tuning Slump WW (liste true/false).",
    )
    parser.add_argument(
        "--l01-slump-models",
        default="gbr",
        help="Modeles Slump L01 (liste).",
    )
    parser.add_argument(
        "--l01-slump-tune-options",
        default="true",
        help="Options tuning Slump L01 (liste true/false).",
    )
    return parser.parse_args(argv)


def _parse_list(value: str) -> list[str]:
    """Parse une liste separee par virgules."""
    return [item.strip() for item in value.split(",") if item.strip()]


def _parse_list_lower(value: str) -> list[str]:
    """Parse une liste en forcant les valeurs en minuscules."""
    return [item.strip().lower() for item in value.split(",") if item.strip()]


def _parse_bool_list(value: str) -> list[str]:
    """Parse une liste booleenne (true/false)."""
    values = _parse_list_lower(value)
    return ["true" if v in {"true", "1", "yes", "y"} else "false" for v in values]


def _parse_int_list(value: str) -> list[int]:
    """Parse une liste d'entiers."""
    return [int(item) for item in _parse_list(value)]


def _default_datasets() -> list[dict]:
    """Liste des scenarios par defaut (old + new)."""
    return [
        {
            "dataset_id": "old",
            "l01_path": "data/L01-Optimisation.xlsx",
            "l01_sheet": "Feuil1 (2)",
            "ww_path": "data/WW-Optimisation.xlsx",
            "ww_sheet": "Feuil3 (3)",
        },
        {
            "dataset_id": "new",
            "l01_path": "data/L01-dataset.xlsx",
            "l01_sheet": None,
            "ww_path": "data/WW-Optimisation.xlsx",
            "ww_sheet": "Feuil3 (3)",
        },
    ]


def _load_datasets(path: str | None) -> list[dict]:
    """Charge la liste des scenarios depuis un JSON (optionnel)."""
    if not path:
        return _default_datasets()

    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if isinstance(data, dict):
        datasets = data.get("datasets")
    elif isinstance(data, list):
        datasets = data
    else:
        raise ValueError("Format de datasets JSON invalide.")

    if not isinstance(datasets, list):
        raise ValueError("Le JSON datasets doit contenir une liste.")

    normalized = []
    for item in datasets:
        if not isinstance(item, dict):
            raise ValueError("Chaque dataset doit etre un objet JSON.")
        dataset_id = item.get("dataset_id") or item.get("id") or item.get("name")
        l01_path = item.get("l01_path") or item.get("l01")
        ww_path = item.get("ww_path") or item.get("ww")
        l01_sheet = item.get("l01_sheet") or item.get("sheet_l01")
        ww_sheet = item.get("ww_sheet") or item.get("sheet_ww")

        if not dataset_id or not l01_path or not ww_path:
            raise ValueError("dataset_id, l01_path et ww_path sont requis.")
        normalized.append(
            {
                "dataset_id": str(dataset_id),
                "l01_path": str(l01_path),
                "l01_sheet": l01_sheet,
                "ww_path": str(ww_path),
                "ww_sheet": ww_sheet,
            }
        )
    return normalized


def _first_sheet(path: str) -> str:
    """Retourne la premiere feuille Excel (openpyxl)."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if not workbook.sheetnames:
            raise ValueError("Aucune feuille Excel trouvee.")
        return workbook.sheetnames[0]
    finally:
        workbook.close()


def _resolve_sheets(datasets: list[dict]) -> list[dict]:
    """Resout les feuilles manquantes pour chaque dataset."""
    resolved = []
    for dataset in datasets:
        dataset_id = dataset["dataset_id"]
        try:
            l01_sheet = dataset.get("l01_sheet")
            ww_sheet = dataset.get("ww_sheet")
            if not l01_sheet:
                l01_sheet = _first_sheet(dataset["l01_path"])
            if not ww_sheet:
                ww_sheet = _first_sheet(dataset["ww_path"])
            dataset["l01_sheet"] = l01_sheet
            dataset["ww_sheet"] = ww_sheet
            resolved.append(dataset)
        except Exception as exc:
            print(f"[{dataset_id}] Erreur sheets: {exc}")
    return resolved


def _build_run_id(cfg: dict) -> str:
    """Construit un identifiant de run stable et lisible."""
    return (
        f"{cfg['dataset_id']}_seed{cfg['seed']}_{cfg['search_mode']}"
        f"_WW{cfg['ww_ucs_model']}-{cfg['ww_ucs_transform']}"
        f"-{cfg['ww_ucs_outliers']}-t{cfg['ww_tune']}"
        f"_L01{cfg['l01_ucs_model']}-{cfg['l01_ucs_transform']}"
        f"-{cfg['l01_ucs_outliers']}-t{cfg['l01_tune']}"
    )


def _iter_run_configs(
    datasets: list[dict],
    seeds: list[int],
    search_modes: list[str],
    ww_ucs_models: list[str],
    ww_ucs_transforms: list[str],
    ww_ucs_outliers: list[str],
    ww_tunes: list[str],
    l01_ucs_models: list[str],
    l01_ucs_transforms: list[str],
    l01_ucs_outliers: list[str],
    l01_tunes: list[str],
    ww_slump_models: list[str],
    ww_slump_tunes: list[str],
    l01_slump_models: list[str],
    l01_slump_tunes: list[str],
) -> list[dict]:
    """Genere la grille complete de configurations."""
    configs = []
    for dataset in datasets:
        for (
            search_mode,
            ww_model,
            ww_transform,
            ww_outlier,
            ww_tune,
            l01_model,
            l01_transform,
            l01_outlier,
            l01_tune,
            ww_slump_model,
            ww_slump_tune,
            l01_slump_model,
            l01_slump_tune,
            seed,
        ) in product(
            search_modes,
            ww_ucs_models,
            ww_ucs_transforms,
            ww_ucs_outliers,
            ww_tunes,
            l01_ucs_models,
            l01_ucs_transforms,
            l01_ucs_outliers,
            l01_tunes,
            ww_slump_models,
            ww_slump_tunes,
            l01_slump_models,
            l01_slump_tunes,
            seeds,
        ):
            cfg = {
                "dataset_id": dataset["dataset_id"],
                "seed": seed,
                "search_mode": search_mode,
                "ww_ucs_model": ww_model,
                "ww_ucs_transform": ww_transform,
                "ww_ucs_outliers": ww_outlier,
                "ww_tune": ww_tune,
                "l01_ucs_model": l01_model,
                "l01_ucs_transform": l01_transform,
                "l01_ucs_outliers": l01_outlier,
                "l01_tune": l01_tune,
                "ww_slump_model": ww_slump_model,
                "ww_slump_tune": ww_slump_tune,
                "l01_slump_model": l01_slump_model,
                "l01_slump_tune": l01_slump_tune,
            }
            cfg["run_id"] = _build_run_id(cfg)
            configs.append(cfg)
    return configs


def _safe_get(metrics: dict, *keys) -> float:
    """Acces securise a un champ de metrics.json."""
    cur = metrics
    for key in keys:
        if not isinstance(cur, dict) or key not in cur:
            return float("nan")
        cur = cur[key]
    try:
        return float(cur)
    except (TypeError, ValueError):
        return float("nan")


def _compute_run_score(row: dict) -> float:
    """Calcule un score simple pour classer les runs.

    Heuristique: priorite au L01 (R2/RMSE), puis penalise le slump.
    Ce score sert uniquement a ordonner les configs, pas a prouver une verite.
    """
    values = [
        row.get("ucs_r2_l01", float("nan")),
        row.get("ucs_rmse_l01", float("nan")),
        row.get("ucs_r2_ww", float("nan")),
        row.get("slump_rmse_overall", float("nan")),
    ]
    if any(math.isnan(v) for v in values):
        return float("nan")
    return (
        row["ucs_r2_l01"]
        - (row["ucs_rmse_l01"] / 1000.0)
        + 0.20 * row["ucs_r2_ww"]
        - 0.05 * (row["slump_rmse_overall"] / 10.0)
    )


def _nanmean(values: list[float]) -> float:
    """Moyenne robuste a NaN."""
    clean = [v for v in values if not math.isnan(v)]
    if not clean:
        return float("nan")
    return float(statistics.mean(clean))


def _nanstd(values: list[float]) -> float:
    """Ecart-type robuste a NaN."""
    clean = [v for v in values if not math.isnan(v)]
    if not clean:
        return float("nan")
    if len(clean) == 1:
        return 0.0
    return float(statistics.pstdev(clean))


def _write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    """Ecrit une liste de dicts en CSV."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _format_config_id(row: dict) -> str:
    """Formate un id court de configuration pour l'affichage."""
    return (
        f"{row['dataset_id']}_{row['search_mode']}"
        f"_WW{row['ww_ucs_model']}-{row['ww_ucs_transform']}"
        f"-{row['ww_ucs_outliers']}-t{row['ww_tune']}"
        f"_L01{row['l01_ucs_model']}-{row['l01_ucs_transform']}"
        f"-{row['l01_ucs_outliers']}-t{row['l01_tune']}"
        f"_SWW{row['ww_slump_model']}-t{row['ww_slump_tune']}"
        f"_SL01{row['l01_slump_model']}-t{row['l01_slump_tune']}"
    )


def _print_top(rows: list[dict], title: str, key: str, reverse: bool) -> None:
    """Affiche un top 5 selon une metrique."""
    print(title)
    filtered = [
        row
        for row in rows
        if not math.isnan(float(row.get(key, float("nan"))))
    ]
    sorted_rows = sorted(
        filtered,
        key=lambda r: float(r.get(key, float("nan"))),
        reverse=reverse,
    )
    for row in sorted_rows[:5]:
        value = row.get(key, float("nan"))
        print(f"  {_format_config_id(row)}: {value}")


def _aggregate_configs(rows: list[dict]) -> list[dict]:
    """Agrege les runs par configuration (moyenne + ecart-type)."""
    grouped: dict[tuple, list[dict]] = {}
    for row in rows:
        key = tuple(row[field] for field in CONFIG_FIELDS)
        grouped.setdefault(key, []).append(row)

    aggregated: list[dict] = []
    for key, items in grouped.items():
        row = {field: value for field, value in zip(CONFIG_FIELDS, key)}
        row["n_runs"] = len(items)
        for metric in AGG_METRICS:
            values = [float(item.get(metric, float("nan"))) for item in items]
            row[f"mean_{metric}"] = _nanmean(values)
            row[f"std_{metric}"] = _nanstd(values)
        mean_score = row.get("mean_run_score", float("nan"))
        std_score = row.get("std_run_score", float("nan"))
        if math.isnan(mean_score):
            row["final_score"] = float("nan")
        else:
            row["final_score"] = mean_score - 0.25 * (
                0.0 if math.isnan(std_score) else std_score
            )
        aggregated.append(row)
    return aggregated


def _collect_run_rows(
    run_configs: list[dict],
    datasets_by_id: dict[str, dict],
    out_dir: Path,
) -> list[dict]:
    """Charge les metrics.json valides pour produire sweep_runs.csv."""
    rows: list[dict] = []
    for cfg in run_configs:
        dataset_id = cfg["dataset_id"]
        dataset_out_dir = out_dir / dataset_id
        run_dir = dataset_out_dir / "runs" / cfg["run_id"]
        status_path = run_dir / "status.json"
        metrics_path = run_dir / "metrics.json"
        if not metrics_path.exists():
            continue
        if status_path.exists():
            status = json.loads(status_path.read_text(encoding="utf-8"))
            if status.get("status") != "ok":
                continue

        metrics = json.loads(metrics_path.read_text(encoding="utf-8"))
        row = {field: cfg[field] for field in CONFIG_FIELDS}
        row.update(
            {
                "dataset_id": dataset_id,
                "run_id": cfg["run_id"],
                "seed": cfg["seed"],
                "search_mode": cfg["search_mode"],
                "ucs_r2_l01": _safe_get(
                    metrics, "UCS", "Tailings", "L01", "r2"
                ),
                "ucs_rmse_l01": _safe_get(
                    metrics, "UCS", "Tailings", "L01", "rmse"
                ),
                "ucs_r2_ww": _safe_get(
                    metrics, "UCS", "Tailings", "WW", "r2"
                ),
                "ucs_rmse_ww": _safe_get(
                    metrics, "UCS", "Tailings", "WW", "rmse"
                ),
                "slump_r2_overall": _safe_get(
                    metrics, "Slump", "overall", "r2"
                ),
                "slump_rmse_overall": _safe_get(
                    metrics, "Slump", "overall", "rmse"
                ),
                "pass_rate_pct": _safe_get(
                    metrics, "optimisation", "pass_rate_pct"
                ),
            }
        )
        row["run_score"] = _compute_run_score(row)
        rows.append(row)
    return rows


def _run_cli(
    cfg: dict,
    dataset: dict,
    args: argparse.Namespace,
    base_out_dir: Path,
) -> bool:
    """Lance un run via subprocess et stocke command.txt + run.log."""
    dataset_out_dir = base_out_dir / dataset["dataset_id"]
    run_dir = dataset_out_dir / "runs" / cfg["run_id"]
    run_dir.mkdir(parents=True, exist_ok=True)
    xlsx_dir = dataset_out_dir / "xlsx"
    xlsx_dir.mkdir(parents=True, exist_ok=True)
    out_xlsx = xlsx_dir / f"{cfg['run_id']}.xlsx"

    cmd = [
        sys.executable,
        "-m",
        "src.cli",
        "--fit-mode",
        "hybrid_by_tailings",
        "--seed",
        str(cfg["seed"]),
        "--run-id",
        cfg["run_id"],
        "--out-dir",
        str(dataset_out_dir),
        "--l01",
        dataset["l01_path"],
        "--ww",
        dataset["ww_path"],
        "--sheet-l01",
        dataset["l01_sheet"],
        "--sheet-ww",
        dataset["ww_sheet"],
        "--slump-min",
        str(args.slump_min),
        "--ucs-min",
        str(args.ucs_min),
        "--n-samples",
        str(args.n_samples),
        "--search-mode",
        cfg["search_mode"],
        "--ww-model",
        cfg["ww_ucs_model"],
        "--ww-ucs-transform",
        cfg["ww_ucs_transform"],
        "--ww-ucs-outliers",
        cfg["ww_ucs_outliers"],
        "--ww-tune",
        cfg["ww_tune"],
        "--l01-model",
        cfg["l01_ucs_model"],
        "--l01-ucs-transform",
        cfg["l01_ucs_transform"],
        "--l01-ucs-outliers",
        cfg["l01_ucs_outliers"],
        "--l01-tune",
        cfg["l01_tune"],
        "--ww-slump-model",
        cfg["ww_slump_model"],
        "--ww-slump-tune",
        cfg["ww_slump_tune"],
        "--l01-slump-model",
        cfg["l01_slump_model"],
        "--l01-slump-tune",
        cfg["l01_slump_tune"],
        "--out",
        str(out_xlsx),
    ]

    (run_dir / "command.txt").write_text(
        "\n".join(cmd), encoding="utf-8"
    )

    try:
        # On capture stdout/stderr dans run.log pour debug.
        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            check=False,
        )
        log_text = result.stdout or ""
        log_text += f"\n[returncode] {result.returncode}\n"
        (run_dir / "run.log").write_text(log_text, encoding="utf-8")
        status = {
            "status": "ok" if result.returncode == 0 else "error",
            "returncode": result.returncode,
        }
    except Exception as exc:
        (run_dir / "run.log").write_text(str(exc), encoding="utf-8")
        status = {"status": "error", "error": str(exc)}

    (run_dir / "status.json").write_text(
        json.dumps(status, indent=2), encoding="utf-8"
    )

    if status.get("status") != "ok":
        return False

    produced = run_dir / "Top_Recipes.xlsx"
    if produced.exists():
        out_xlsx.parent.mkdir(parents=True, exist_ok=True)
        out_xlsx.write_bytes(produced.read_bytes())
    return True


def main(argv: list[str] | None = None) -> int:
    """Point d'entree principal du sweep."""
    args = _parse_args(argv)
    if args.jobs and args.jobs != 1:
        print("jobs>1 non supporte, execution sequentielle.")

    try:
        datasets = _load_datasets(args.datasets)
    except Exception as exc:
        print(f"Erreur datasets: {exc}")
        return 1

    if args.only_dataset:
        target_id = args.only_dataset.strip().lower()
        datasets = [
            ds
            for ds in datasets
            if str(ds.get("dataset_id", "")).lower() == target_id
        ]
        if not datasets:
            print(f"Aucun dataset_id correspondant a {args.only_dataset}.")
            return 1

    seeds = _parse_int_list(args.seeds)
    search_modes = _parse_list_lower(args.search_modes)
    ww_ucs_models = _parse_list_lower(args.ww_ucs_models)
    ww_ucs_transforms = _parse_list_lower(args.ww_ucs_transforms)
    ww_ucs_outliers = _parse_list_lower(args.ww_ucs_outliers)
    ww_tunes = _parse_bool_list(args.ww_tune_options)
    l01_ucs_models = _parse_list_lower(args.l01_ucs_models)
    l01_ucs_transforms = _parse_list_lower(args.l01_ucs_transforms)
    l01_ucs_outliers = _parse_list_lower(args.l01_ucs_outliers)
    l01_tunes = _parse_bool_list(args.l01_tune_options)
    ww_slump_models = _parse_list_lower(args.ww_slump_models)
    ww_slump_tunes = _parse_bool_list(args.ww_slump_tune_options)
    l01_slump_models = _parse_list_lower(args.l01_slump_models)
    l01_slump_tunes = _parse_bool_list(args.l01_slump_tune_options)

    configs = _iter_run_configs(
        datasets,
        seeds,
        search_modes,
        ww_ucs_models,
        ww_ucs_transforms,
        ww_ucs_outliers,
        ww_tunes,
        l01_ucs_models,
        l01_ucs_transforms,
        l01_ucs_outliers,
        l01_tunes,
        ww_slump_models,
        ww_slump_tunes,
        l01_slump_models,
        l01_slump_tunes,
    )

    if args.max_runs and args.max_runs > 0:
        configs = configs[: args.max_runs]

    if args.dry_run:
        print(f"Nombre de runs prevus: {len(configs)}")
        return 0

    datasets = _resolve_sheets(datasets)
    if not datasets:
        print("Aucun dataset valide.")
        return 1
    datasets_by_id = {d["dataset_id"]: d for d in datasets}

    base_out_dir = Path(args.out_dir)
    start_time = time.monotonic()
    budget_seconds = max(args.time_budget_min, 0) * 60.0
    budget_exceeded = False
    executed = 0
    skipped = 0

    for cfg in configs:
        if budget_seconds:
            elapsed = time.monotonic() - start_time
            if elapsed >= budget_seconds and not budget_exceeded:
                # Budget temps "souple": on informe, mais on continue les runs.
                print(
                    "Budget temps atteint (limite souple), poursuite des runs."
                )
                budget_exceeded = True
        dataset_id = cfg["dataset_id"]
        dataset = datasets_by_id.get(dataset_id)
        if not dataset:
            print(f"[{dataset_id}] Dataset inconnu, run ignore.")
            skipped += 1
            continue

        dataset_out_dir = base_out_dir / dataset_id
        run_dir = dataset_out_dir / "runs" / cfg["run_id"]
        metrics_path = run_dir / "metrics.json"
        if args.resume and metrics_path.exists():
            # Resume: on saute si metrics.json existe deja.
            skipped += 1
            continue

        ok = _run_cli(cfg, dataset, args, base_out_dir)
        if ok:
            executed += 1
        else:
            skipped += 1

    all_rows: list[dict] = []
    for dataset in datasets:
        dataset_id = dataset["dataset_id"]
        dataset_out_dir = base_out_dir / dataset_id
        dataset_configs = [
            cfg for cfg in configs if cfg["dataset_id"] == dataset_id
        ]
        run_rows = _collect_run_rows(dataset_configs, datasets_by_id, base_out_dir)
        if run_rows:
            _write_csv(dataset_out_dir / "sweep_runs.csv", run_rows, RUN_FIELDS)
            all_rows.extend(run_rows)
            config_rows = _aggregate_configs(run_rows)
            agg_fields = (
                CONFIG_FIELDS
                + ["n_runs"]
                + [f"mean_{m}" for m in AGG_METRICS]
                + [f"std_{m}" for m in AGG_METRICS]
                + ["final_score"]
            )
            _write_csv(
                dataset_out_dir / "sweep_configs.csv", config_rows, agg_fields
            )
            print(f"[{dataset_id}] Top 5 par final_score")
            _print_top(config_rows, "Top 5 par final_score", "final_score", True)
            _print_top(
                config_rows,
                "Top 5 par mean_ucs_r2_l01",
                "mean_ucs_r2_l01",
                True,
            )
            _print_top(
                config_rows,
                "Top 5 par mean_ucs_rmse_l01 (plus bas)",
                "mean_ucs_rmse_l01",
                False,
            )
            _print_top(
                config_rows,
                "Top 5 par mean_ucs_r2_ww",
                "mean_ucs_r2_ww",
                True,
            )
        else:
            print(f"[{dataset_id}] Aucun run valide.")

    if all_rows:
        _write_csv(base_out_dir / "sweep_runs.csv", all_rows, RUN_FIELDS)
        config_rows = _aggregate_configs(all_rows)
        agg_fields = (
            CONFIG_FIELDS
            + ["n_runs"]
            + [f"mean_{m}" for m in AGG_METRICS]
            + [f"std_{m}" for m in AGG_METRICS]
            + ["final_score"]
        )
        _write_csv(base_out_dir / "sweep_configs.csv", config_rows, agg_fields)
        print("[ALL] Top 5 par final_score")
        _print_top(config_rows, "Top 5 par final_score", "final_score", True)
        _print_top(
            config_rows,
            "Top 5 par mean_ucs_r2_l01",
            "mean_ucs_r2_l01",
            True,
        )
        _print_top(
            config_rows,
            "Top 5 par mean_ucs_rmse_l01 (plus bas)",
            "mean_ucs_rmse_l01",
            False,
        )
        _print_top(
            config_rows,
            "Top 5 par mean_ucs_r2_ww",
            "mean_ucs_r2_ww",
            True,
        )

    print(
        f"Sweep termine. Executes={executed} Skipped={skipped} "
        f"Total={len(configs)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
