"""Interface en ligne de commande.

Ce module orchestre tout le pipeline :
- lecture des donnees Excel (L01 + WW)
- nettoyage minimal et normalisation des colonnes requises
- entrainement des modeles (combined / by_tailings / hybrid_by_tailings)
- evaluation (CV) et export des metriques
- optimisation Monte-Carlo et export Excel final
- sauvegarde des modeles et metadonnees pour la reproductibilite
"""

from __future__ import annotations

import argparse
import json
import random
import sys
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .config import DEFAULTS, TARGET_SLUMP, TARGET_UCS
from .features import (
    coerce_numeric,
    infer_feature_columns,
    make_training_frames,
    split_features_target,
)
from .io_data import read_excel_file, validate_dataframe
from .optimize import export_top_recipes, optimize_recipes
from .schema import clean_dataframe, standardize_required_columns
from .train import (
    build_estimator,
    cross_validate_report,
    cross_validate_report_hybrid,
    filter_ucs_outliers,
    fit_models_hybrid,
    get_model_meta,
    tune_estimator,
)


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Construit l'interface CLI et parse les arguments utilisateur."""
    parser = argparse.ArgumentParser(
        description=(
            "Entraine les modeles Slump et UCS et optimise les recettes CPB."
        )
    )
    parser.add_argument(
        "--l01", required=True, help="Chemin du fichier Excel L01."
    )
    parser.add_argument(
        "--ww", required=True, help="Chemin du fichier Excel WW."
    )
    parser.add_argument(
        "--sheet-l01",
        default=DEFAULTS.sheet_l01,
        help="Nom de feuille pour L01 (defaut : premiere feuille).",
    )
    parser.add_argument(
        "--sheet-ww",
        default=DEFAULTS.sheet_ww,
        help="Nom de feuille pour WW (defaut : premiere feuille).",
    )
    parser.add_argument(
        "--print-columns",
        action="store_true",
        help="Affiche les colonnes detectees puis quitte.",
    )
    parser.add_argument(
        "--slump-min",
        type=float,
        default=DEFAULTS.slump_min,
        help="Seuil minimum de Slump (mm) predit.",
    )
    parser.add_argument(
        "--ucs-min",
        type=float,
        default=DEFAULTS.ucs_min,
        help="Seuil minimum de UCS28d (kPa) predit.",
    )
    parser.add_argument(
        "--n-samples",
        type=int,
        default=DEFAULTS.n_samples,
        help="Nombre de recettes aleatoires par groupe Tailings/Binder.",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=42,
        help="Graine aleatoire pour la reproductibilite.",
    )
    parser.add_argument(
        "--fit-mode",
        choices=["combined", "by_tailings", "hybrid_by_tailings"],
        default="combined",
        help="Mode d'entrainement des modeles.",
    )
    parser.add_argument(
        "--model",
        default="gbr",
        choices=["gbr", "rf", "et", "hgb", "svr", "enet"],
        help="Modele a utiliser pour Slump et UCS.",
    )
    parser.add_argument(
        "--target-transform-ucs",
        choices=["none", "log"],
        default="none",
        help="Transformation de la cible UCS (none ou log).",
    )
    parser.add_argument(
        "--outliers-ucs",
        choices=["none", "iqr", "zscore"],
        default="none",
        help="Gestion des outliers UCS (none, iqr ou zscore).",
    )
    parser.add_argument(
        "--tune",
        action="store_true",
        help="Active un tuning simple des hyperparametres.",
    )
    parser.add_argument(
        "--ww-model",
        choices=["gbr", "rf", "et", "hgb", "svr", "enet"],
        default=None,
        help="Modele UCS pour WW (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--l01-model",
        choices=["gbr", "rf", "et", "hgb", "svr", "enet"],
        default=None,
        help="Modele UCS pour L01 (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--ww-ucs-transform",
        choices=["none", "log"],
        default=None,
        help="Transformation UCS pour WW (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--l01-ucs-transform",
        choices=["none", "log"],
        default=None,
        help="Transformation UCS pour L01 (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--ww-ucs-outliers",
        choices=["none", "iqr", "zscore"],
        default=None,
        help="Outliers UCS pour WW (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--ww-ucs-fixed-params",
        default="",
        help="JSON de parametres fixes UCS pour WW (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--l01-ucs-outliers",
        choices=["none", "iqr", "zscore"],
        default=None,
        help="Outliers UCS pour L01 (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--l01-ucs-fixed-params",
        default="",
        help="JSON de parametres fixes UCS pour L01 (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--ww-tune",
        choices=["true", "false"],
        default=None,
        help="Tuning UCS pour WW (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--l01-tune",
        choices=["true", "false"],
        default=None,
        help="Tuning UCS pour L01 (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--ww-slump-model",
        choices=["gbr", "rf", "et", "hgb", "svr", "enet"],
        default=None,
        help="Modele Slump pour WW (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--l01-slump-model",
        choices=["gbr", "rf", "et", "hgb", "svr", "enet"],
        default=None,
        help="Modele Slump pour L01 (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--ww-slump-tune",
        choices=["true", "false"],
        default=None,
        help="Tuning Slump pour WW (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--ww-slump-fixed-params",
        default="",
        help="JSON de parametres fixes Slump pour WW (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--l01-slump-tune",
        choices=["true", "false"],
        default=None,
        help="Tuning Slump pour L01 (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--l01-slump-fixed-params",
        default="",
        help="JSON de parametres fixes Slump pour L01 (hybrid_by_tailings).",
    )
    parser.add_argument(
        "--search-mode",
        choices=["uniform", "bootstrap"],
        default="uniform",
        help="Mode de recherche pour l'echantillonnage.",
    )
    parser.add_argument(
        "--export-all",
        action="store_true",
        help="Exporte aussi tous les candidats avant filtrage.",
    )
    parser.add_argument(
        "--out",
        default="Top_Recipes.xlsx",
        help="Chemin du fichier Excel de sortie.",
    )
    parser.add_argument(
        "--save-models",
        choices=["true", "false"],
        default="true",
        help="Sauvegarde les modeles entraines.",
    )
    parser.add_argument(
        "--models-dir",
        default="",
        help="Dossier de sauvegarde des modeles (defaut: outputs/models/<run-id>).",
    )
    parser.add_argument(
        "--out-dir",
        default="outputs",
        help="Dossier de sortie principal.",
    )
    parser.add_argument(
        "--run-id",
        default="",
        help="Identifiant de run (optionnel).",
    )
    return parser.parse_args(argv)


def _to_jsonable(value):
    """Convertit les types numpy en types JSON serialisables."""
    if isinstance(value, dict):
        return {k: _to_jsonable(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(v) for v in value]
    if isinstance(value, np.generic):
        return value.item()
    return value


def _parse_bool(value: str | None, fallback: bool = False) -> bool:
    """Parse une valeur 'true/false' optionnelle."""
    if value is None:
        return fallback
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def _resolve(value: str | None, fallback: str) -> str:
    """Retourne value si renseigne, sinon fallback."""
    return value if value else fallback


def _load_fixed_params(path: str | None) -> dict | None:
    """Charge un JSON de parametres fixes pour un pipeline sklearn."""
    if not path:
        return None
    param_path = Path(path)
    if not param_path.exists():
        raise FileNotFoundError(f"Fichier params introuvable : {param_path}")
    data = json.loads(param_path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError(
            f"Params fixes invalides dans {param_path} (dict attendu)."
        )
    return data


def _first_sheet(path: str | Path) -> str:
    """Recupere la premiere feuille d'un fichier Excel."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if not workbook.sheetnames:
            raise ValueError(f"Aucune feuille trouvee dans {path}")
        return workbook.sheetnames[0]
    finally:
        workbook.close()


def _build_by_tailings_config(args: argparse.Namespace) -> dict:
    """Construit une config identique WW/L01 (mode by_tailings)."""
    return {
        "WW": {
            "slump": {"model": args.model, "tune": args.tune},
            "ucs": {
                "model": args.model,
                "transform": args.target_transform_ucs,
                "outliers": args.outliers_ucs,
                "tune": args.tune,
            },
        },
        "L01": {
            "slump": {"model": args.model, "tune": args.tune},
            "ucs": {
                "model": args.model,
                "transform": args.target_transform_ucs,
                "outliers": args.outliers_ucs,
                "tune": args.tune,
            },
        },
    }


def _build_hybrid_config(args: argparse.Namespace) -> dict:
    """Construit une config par tailings (mode hybrid_by_tailings)."""
    ww_model = _resolve(args.ww_model, args.model)
    l01_model = _resolve(args.l01_model, args.model)
    ww_slump_model = _resolve(args.ww_slump_model, args.model)
    l01_slump_model = _resolve(args.l01_slump_model, args.model)

    ww_ucs_fixed = _load_fixed_params(args.ww_ucs_fixed_params)
    l01_ucs_fixed = _load_fixed_params(args.l01_ucs_fixed_params)
    ww_slump_fixed = _load_fixed_params(args.ww_slump_fixed_params)
    l01_slump_fixed = _load_fixed_params(args.l01_slump_fixed_params)

    ww_tune = _parse_bool(args.ww_tune, args.tune)
    l01_tune = _parse_bool(args.l01_tune, args.tune)
    ww_slump_tune = _parse_bool(args.ww_slump_tune, ww_tune)
    l01_slump_tune = _parse_bool(args.l01_slump_tune, l01_tune)

    if ww_ucs_fixed:
        ww_tune = False
    if l01_ucs_fixed:
        l01_tune = False
    if ww_slump_fixed:
        ww_slump_tune = False
    if l01_slump_fixed:
        l01_slump_tune = False

    return {
        "WW": {
            "slump": {
                "model": ww_slump_model,
                "tune": ww_slump_tune,
                "fixed_params": ww_slump_fixed,
            },
            "ucs": {
                "model": ww_model,
                "transform": _resolve(
                    args.ww_ucs_transform, args.target_transform_ucs
                ),
                "outliers": _resolve(
                    args.ww_ucs_outliers, args.outliers_ucs
                ),
                "tune": ww_tune,
                "fixed_params": ww_ucs_fixed,
            },
        },
        "L01": {
            "slump": {
                "model": l01_slump_model,
                "tune": l01_slump_tune,
                "fixed_params": l01_slump_fixed,
            },
            "ucs": {
                "model": l01_model,
                "transform": _resolve(
                    args.l01_ucs_transform, args.target_transform_ucs
                ),
                "outliers": _resolve(
                    args.l01_ucs_outliers, args.outliers_ucs
                ),
                "tune": l01_tune,
                "fixed_params": l01_ucs_fixed,
            },
        },
    }


def main(argv: list[str] | None = None) -> int:
    """Point d'entree principal, utilisable par import ou via CLI.

    Rappels importants:
        - Reproductibilite: fixer --seed + reutiliser des fixed params si besoin.
        - fit_mode:
            * combined: un modele global
            * by_tailings: meme config pour WW et L01
            * hybrid_by_tailings: config distincte par tailings
        - search-mode:
            * uniform: tirage min/max
            * bootstrap: tirage sur valeurs observees
        - outputs:
            * sans --run-id: ecrit dans --out-dir
            * avec --run-id: ecrit dans --out-dir/runs/<run_id>/
    """
    args = _parse_args(argv)
    random.seed(args.seed)
    np.random.seed(args.seed)
    try:
        # Auto-detection des feuilles si non precisees.
        sheet_l01 = args.sheet_l01 or _first_sheet(args.l01)
        sheet_ww = args.sheet_ww or _first_sheet(args.ww)

        # Lecture + nettoyage minimal par fichier, puis normalisation des colonnes requises.
        l01_df = read_excel_file(args.l01, sheet_l01)
        l01_df["Tailings"] = "L01"
        # IMPORTANT: on nettoie puis on normalise uniquement les colonnes requises.
        # Le schema NEW peut avoir d'autres colonnes numeriques, on les conserve.
        l01_df = clean_dataframe(l01_df)
        l01_df = standardize_required_columns(l01_df)

        ww_df = read_excel_file(args.ww, sheet_ww)
        ww_df["Tailings"] = "WW"
        ww_df = clean_dataframe(ww_df)
        ww_df = standardize_required_columns(ww_df)

        if args.print_columns:
            # Mode debug: on affiche les colonnes detectees et on quitte.
            l01_cat, l01_num = infer_feature_columns(
                l01_df, target_cols=[TARGET_SLUMP, TARGET_UCS]
            )
            ww_cat, ww_num = infer_feature_columns(
                ww_df, target_cols=[TARGET_SLUMP, TARGET_UCS]
            )
            print(f"L01 columns: {list(l01_df.columns)}")
            print(f"WW columns: {list(ww_df.columns)}")
            print(f"L01 numeric columns used: {l01_num}")
            print(f"L01 categorical columns used: {l01_cat}")
            print(f"WW numeric columns used: {ww_num}")
            print(f"WW categorical columns used: {ww_cat}")
            return 0

        df = pd.concat([l01_df, ww_df], ignore_index=True, sort=False)
        validate_dataframe(df)
        # Option A: on separe les lignes valides par cible (pas de drop global).
        df_slump, df_ucs = make_training_frames(df)

        metrics = {}
        models_by_tailings = None
        best_params = {}

        features_meta = {}
        if args.fit_mode == "combined":
            # Un seul modele Slump et un seul modele UCS pour tout le dataset.
            if args.outliers_ucs != "none":
                df_ucs = filter_ucs_outliers(df_ucs, args.outliers_ucs)

            slump_cat, slump_num = infer_feature_columns(
                df_slump, target_cols=[TARGET_SLUMP, TARGET_UCS]
            )
            X_slump, y_slump = split_features_target(
                df_slump, TARGET_SLUMP, slump_cat, slump_num
            )
            X_slump = coerce_numeric(X_slump, slump_num)
            y_slump = coerce_numeric(y_slump.to_frame(), [TARGET_SLUMP])[
                TARGET_SLUMP
            ]

            ucs_cat, ucs_num = infer_feature_columns(
                df_ucs, target_cols=[TARGET_SLUMP, TARGET_UCS]
            )
            X_ucs, y_ucs = split_features_target(
                df_ucs, TARGET_UCS, ucs_cat, ucs_num
            )
            X_ucs = coerce_numeric(X_ucs, ucs_num)
            y_ucs = coerce_numeric(y_ucs.to_frame(), [TARGET_UCS])[TARGET_UCS]

            features_meta = {
                "slump": {"categorical": slump_cat, "numeric": slump_num},
                "ucs": {"categorical": ucs_cat, "numeric": ucs_num},
            }

            slump_mask = y_slump.notna()
            X_slump = X_slump.loc[slump_mask].copy()
            y_slump = y_slump.loc[slump_mask].copy()
            df_slump = df_slump.loc[slump_mask].copy()

            ucs_mask = y_ucs.notna()
            X_ucs = X_ucs.loc[ucs_mask].copy()
            y_ucs = y_ucs.loc[ucs_mask].copy()
            df_ucs = df_ucs.loc[ucs_mask].copy()

            if args.target_transform_ucs == "log":
                if (y_ucs <= -1).any():
                    raise ValueError(
                        "UCS contient des valeurs <= -1, log1p impossible."
                    )

            slump_estimator = build_estimator(
                args.model,
                numeric_cols=slump_num,
                categorical_cols=slump_cat,
                random_state=args.seed,
            )
            ucs_estimator = build_estimator(
                args.model,
                target_transform=args.target_transform_ucs,
                numeric_cols=ucs_num,
                categorical_cols=ucs_cat,
                random_state=args.seed,
            )

            best_params = {"slump": {}, "ucs": {}}
            if args.tune:
                # Tuning optionnel (RandomizedSearchCV).
                slump_estimator, best_params["slump"] = tune_estimator(
                    slump_estimator,
                    args.model,
                    X_slump,
                    y_slump,
                    random_state=args.seed,
                )
                ucs_estimator, best_params["ucs"] = tune_estimator(
                    ucs_estimator,
                    args.model,
                    X_ucs,
                    y_ucs,
                    random_state=args.seed,
                )

            metrics = {
                "Slump": cross_validate_report(
                    slump_estimator,
                    X_slump,
                    y_slump,
                    groups_df=df_slump,
                    random_state=args.seed,
                ),
                "UCS": cross_validate_report(
                    ucs_estimator,
                    X_ucs,
                    y_ucs,
                    groups_df=df_ucs,
                    random_state=args.seed,
                ),
            }

            slump_estimator.fit(X_slump, y_slump)
            ucs_estimator.fit(X_ucs, y_ucs)
        else:
            # Mode par tailings: config globale ou hybride (WW/L01 differents).
            if args.fit_mode == "by_tailings":
                config = _build_by_tailings_config(args)
            else:
                config = _build_hybrid_config(args)

            models_by_tailings = fit_models_hybrid(
                df_slump, df_ucs, config, random_state=args.seed
            )
            best_params = {
                tail: models_by_tailings[tail].get("best_params", {})
                for tail in models_by_tailings
            }
            features_meta = {
                tail: models_by_tailings[tail].get("features", {})
                for tail in models_by_tailings
            }
            outliers_map = {
                tail: config[tail]["ucs"]["outliers"] for tail in config
            }

            metrics = {
                "Slump": cross_validate_report_hybrid(
                    models_by_tailings,
                    df_slump,
                    TARGET_SLUMP,
                    random_state=args.seed,
                ),
                "UCS": cross_validate_report_hybrid(
                    models_by_tailings,
                    df_ucs,
                    TARGET_UCS,
                    outliers_by_tailings=outliers_map,
                    random_state=args.seed,
                ),
            }

        base_out_dir = Path(args.out_dir)
        run_dir = None
        if args.run_id:
            # Mode run: toutes les sorties sont isolees dans outputs/runs/<run_id>/.
            run_dir = base_out_dir / "runs" / args.run_id
            run_dir.mkdir(parents=True, exist_ok=True)
            out_path = run_dir / "Top_Recipes.xlsx"
            metrics_path = run_dir / "metrics.json"
            best_params_path = run_dir / "best_params.json"
            export_dir = run_dir
        else:
            # Mode standard: Top_Recipes.xlsx et metrics.json sous --out-dir.
            out_path = Path(args.out)
            if not out_path.is_absolute() and out_path.parent == Path("."):
                out_path = base_out_dir / out_path
            base_out_dir.mkdir(parents=True, exist_ok=True)
            metrics_path = base_out_dir / "metrics.json"
            best_params_path = base_out_dir / "best_params.json"
            export_dir = base_out_dir

        tuning_active = args.tune
        if args.fit_mode in {"by_tailings", "hybrid_by_tailings"}:
            tuning_active = any(
                config[tail][target].get("tune", False)
                for tail in config
                for target in ["slump", "ucs"]
            )

        fixed_params_active = False
        if args.fit_mode in {"by_tailings", "hybrid_by_tailings"}:
            fixed_params_active = any(
                config[tail][target].get("fixed_params")
                for tail in config
                for target in ["slump", "ucs"]
            )

        if tuning_active or fixed_params_active:
            # On enregistre best_params si tuning OU fixed params (reproductibilite).
            best_params_path.write_text(
                json.dumps(_to_jsonable(best_params), indent=2),
                encoding="utf-8",
            )

        if args.fit_mode == "combined":
            recipes, stats = optimize_recipes(
                slump_estimator,
                ucs_estimator,
                df,
                slump_min=args.slump_min,
                ucs_min=args.ucs_min,
                n_samples=args.n_samples,
                search_mode=args.search_mode,
                export_all=args.export_all,
                export_dir=export_dir,
                fit_mode=args.fit_mode,
                seed=args.seed,
            )
        else:
            recipes, stats = optimize_recipes(
                None,
                None,
                df,
                slump_min=args.slump_min,
                ucs_min=args.ucs_min,
                n_samples=args.n_samples,
                search_mode=args.search_mode,
                export_all=args.export_all,
                export_dir=export_dir,
                fit_mode=args.fit_mode,
                models_by_tailings=models_by_tailings,
                seed=args.seed,
            )
        export_top_recipes(recipes, out_path)

        # Metriques d'optimisation (filtrage Monte-Carlo).
        metrics["optimisation"] = {
            "slump_min": args.slump_min,
            "ucs_min": args.ucs_min,
            "n_samples": args.n_samples,
            "pass_rate_pct": stats.get("pass_rate_pct", 0.0),
            "search_mode": args.search_mode,
        }
        meta_models = {}
        if args.fit_mode == "combined":
            meta_models = {
                "slump": {
                    **get_model_meta(slump_estimator, args.model),
                    "best_params": best_params.get("slump", {}),
                },
                "ucs": {
                    **get_model_meta(ucs_estimator, args.model),
                    "best_params": best_params.get("ucs", {}),
                    "transform": args.target_transform_ucs,
                    "outliers": args.outliers_ucs,
                },
            }
        else:
            # Meta modele par tailings pour le mode hybride.
            meta_models = {}
            for tail in ["WW", "L01"]:
                if tail not in models_by_tailings:
                    continue
                tail_models = models_by_tailings[tail]
                meta_models[tail] = {
                    "slump": {
                        **get_model_meta(
                            tail_models["slump_pipe"],
                            config[tail]["slump"]["model"],
                        ),
                        "best_params": tail_models.get("best_params", {}).get(
                            "slump", {}
                        ),
                        "fixed_params": config[tail]["slump"].get(
                            "fixed_params"
                        )
                        or {},
                        "tune": config[tail]["slump"].get("tune", False),
                    },
                    "ucs": {
                        **get_model_meta(
                            tail_models["ucs_pipe"],
                            config[tail]["ucs"]["model"],
                        ),
                        "best_params": tail_models.get("best_params", {}).get(
                            "ucs", {}
                        ),
                        "fixed_params": config[tail]["ucs"].get(
                            "fixed_params"
                        )
                        or {},
                        "transform": config[tail]["ucs"].get("transform", "none"),
                        "outliers": config[tail]["ucs"].get("outliers", "none"),
                        "tune": config[tail]["ucs"].get("tune", False),
                    },
                }

        metrics["meta"] = {
            "out_dir": str(base_out_dir),
            "run_id": args.run_id,
            "fit_mode": args.fit_mode,
            "models": meta_models,
            "features": features_meta,
            "target_transform_ucs": args.target_transform_ucs,
            "outliers_ucs": args.outliers_ucs,
            "tune": tuning_active,
            "seed": args.seed,
        }

        metrics_path.write_text(
            json.dumps(_to_jsonable(metrics), indent=2), encoding="utf-8"
        )

        save_models = _parse_bool(args.save_models, True)
        if save_models:
            # Sauvegarde des pipelines sklearn (joblib) + metadata pour reproduire.
            if args.models_dir:
                models_dir = Path(args.models_dir)
            else:
                suffix = args.run_id if args.run_id else "latest"
                models_dir = base_out_dir / "models" / suffix
            models_dir.mkdir(parents=True, exist_ok=True)

            if args.fit_mode in {"by_tailings", "hybrid_by_tailings"}:
                if models_by_tailings:
                    if "WW" in models_by_tailings:
                        joblib.dump(
                            models_by_tailings["WW"]["ucs_pipe"],
                            models_dir / "ww_ucs.joblib",
                        )
                        joblib.dump(
                            models_by_tailings["WW"]["slump_pipe"],
                            models_dir / "ww_slump.joblib",
                        )
                    if "L01" in models_by_tailings:
                        joblib.dump(
                            models_by_tailings["L01"]["ucs_pipe"],
                            models_dir / "l01_ucs.joblib",
                        )
                        joblib.dump(
                            models_by_tailings["L01"]["slump_pipe"],
                            models_dir / "l01_slump.joblib",
                        )
            else:
                joblib.dump(slump_estimator, models_dir / "slump.joblib")
                joblib.dump(ucs_estimator, models_dir / "ucs.joblib")

            if metrics_path.exists():
                (models_dir / "metrics.json").write_text(
                    metrics_path.read_text(encoding="utf-8"),
                    encoding="utf-8",
                )
            if best_params_path.exists():
                (models_dir / "best_params.json").write_text(
                    best_params_path.read_text(encoding="utf-8"),
                    encoding="utf-8",
                )

            # Metadata pour retracer le run (sources, config, features, etc.).
            metadata = {
                "run_id": args.run_id,
                "seed": args.seed,
                "fit_mode": args.fit_mode,
                "datasets": {
                    "l01": {"path": args.l01, "sheet": sheet_l01},
                    "ww": {"path": args.ww, "sheet": sheet_ww},
                },
                "models": meta_models,
                "features": features_meta,
                "metrics_path": str(metrics_path),
                "best_params_path": str(best_params_path),
            }
            (models_dir / "metadata.json").write_text(
                json.dumps(_to_jsonable(metadata), indent=2),
                encoding="utf-8",
            )
    except Exception as exc:
        print(f"Erreur : {exc}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
