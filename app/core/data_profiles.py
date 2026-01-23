"""Profil des donnees pour warnings et distributions de sampling."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import json

import numpy as np
import pandas as pd
try:
    import streamlit as st
except Exception:  # pragma: no cover - fallback pour tests sans Streamlit
    class _DummyStreamlit:
        def cache_data(self, *args, **kwargs):
            def decorator(func):
                return func
            return decorator

    st = _DummyStreamlit()

from src.schema import clean_dataframe, standardize_required_columns

DATASETS = {
    "WW": {
        "path": "data/WW-Optimisation.xlsx",
        "sheet": "Feuil3 (3)",
        "tailings": "WW",
    },
    "L01_OLD": {
        "path": "data/L01-Optimisation.xlsx",
        "sheet": "Feuil1 (2)",
        "tailings": "L01",
    },
    "L01_NEW": {
        "path": "data/L01-dataset.xlsx",
        "sheet": None,
        "tailings": "L01",
    },
}

PROFILE_PATH = Path("app/assets/data_profiles.json")


@dataclass
class DataProfile:
    """Profil de donnees avec stats simples par colonne numerique."""

    df: pd.DataFrame
    numeric_stats: dict
    categorical_values: dict
    sample_values: dict
    synthetic: bool = False
    bootstrap_ready: bool = False


def _first_sheet(path: str | Path) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if not wb.sheetnames:
            raise ValueError("Aucune feuille detectee.")
        return wb.sheetnames[0]
    finally:
        wb.close()


def _compute_stats(df: pd.DataFrame) -> dict:
    stats = {}
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            series = pd.to_numeric(df[col], errors="coerce")
            if series.notna().any():
                q05 = float(series.quantile(0.05))
                q95 = float(series.quantile(0.95))
                q25 = float(series.quantile(0.25))
                q75 = float(series.quantile(0.75))
                iqr = float(q75 - q25)
            else:
                q05 = q95 = q25 = q75 = iqr = np.nan
            stats[col] = {
                "min": float(series.min(skipna=True)) if series.notna().any() else np.nan,
                "max": float(series.max(skipna=True)) if series.notna().any() else np.nan,
                "mean": float(series.mean(skipna=True)) if series.notna().any() else np.nan,
                "std": float(series.std(skipna=True)) if series.notna().any() else np.nan,
                "p05": q05,
                "p95": q95,
                "q1": q25,
                "q3": q75,
                "iqr": iqr,
            }
    return stats


@st.cache_data(show_spinner=False)
def _load_profiles_json() -> dict:
    """Charge les profils pre-calcules si disponibles (mode cloud)."""
    if not PROFILE_PATH.exists():
        return {}
    try:
        return json.loads(PROFILE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _build_synthetic_df(
    numeric_stats: dict,
    categorical_values: dict,
    sample_values: dict,
    tailings_default: str,
) -> pd.DataFrame:
    """Cree un DataFrame synthetique a partir des stats et d'echantillons.

    Si sample_values est present, on construit un df plus realiste pour
    permettre un bootstrap approximatif sans donnees brutes.
    """
    numeric_cols = list(numeric_stats.keys())
    if not numeric_cols:
        return pd.DataFrame()

    binders = categorical_values.get("Binder") or ["GUL"]
    tailings_list = categorical_values.get("Tailings") or [tailings_default]

    base_len = 2
    if sample_values:
        lengths = [len(v) for v in sample_values.values() if v]
        if lengths:
            base_len = max(lengths)

    data = {}
    for col in numeric_cols:
        values = list(sample_values.get(col, []))
        if values:
            if len(values) < base_len:
                values = values + [np.nan] * (base_len - len(values))
        else:
            min_val = numeric_stats.get(col, {}).get("min")
            max_val = numeric_stats.get(col, {}).get("max")
            values = [min_val, max_val] + [np.nan] * (base_len - 2)
        data[col] = values[:base_len]

    base_df = pd.DataFrame(data)
    rows = []
    for tail in tailings_list:
        for binder in binders:
            df_copy = base_df.copy()
            df_copy["Tailings"] = tail
            df_copy["Binder"] = binder
            rows.append(df_copy)

    df = pd.concat(rows, ignore_index=True)
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


@st.cache_data(show_spinner=False)
def load_profile(dataset_key: str) -> DataProfile | None:
    """Charge un profil de donnees pour un dataset.

    Notes:
        - Utilise le nettoyage minimal + normalisation des colonnes requises.
        - Ajoute/force la colonne Tailings pour coherer les predictions.
    """

    cfg = DATASETS.get(dataset_key)
    if not cfg:
        return None

    path = Path(cfg["path"])
    if path.exists():
        sheet = cfg["sheet"]
        if sheet is None:
            sheet = _first_sheet(path)

        try:
            df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
            df["Tailings"] = cfg["tailings"]
            df = clean_dataframe(df)
            df = standardize_required_columns(df)
        except Exception:
            df = None

        if df is not None:
            numeric_stats = _compute_stats(df)
            categorical_values = {}
            for col in df.columns:
                if col in {"Binder", "Tailings"}:
                    values = sorted({str(v) for v in df[col].dropna().unique()})
                    categorical_values[col] = values
            return DataProfile(
                df=df,
                numeric_stats=numeric_stats,
                categorical_values=categorical_values,
                sample_values={},
                synthetic=False,
                bootstrap_ready=True,
            )

    # Fallback: profil pre-calcule (utile en mode cloud sans data/)
    profiles = _load_profiles_json()
    profile = profiles.get(dataset_key)
    if not profile:
        return None

    numeric_stats = profile.get("numeric_stats", {})
    categorical_values = profile.get("categorical_values", {})
    sample_values = profile.get("sample_values", {})
    df_synth = _build_synthetic_df(
        numeric_stats, categorical_values, sample_values, cfg["tailings"]
    )
    bootstrap_ready = bool(sample_values)

    return DataProfile(
        df=df_synth,
        numeric_stats=numeric_stats,
        categorical_values=categorical_values,
        sample_values=sample_values,
        synthetic=True,
        bootstrap_ready=bootstrap_ready,
    )


def warn_out_of_profile(profile: DataProfile, inputs: dict) -> list[str]:
    """Retourne une liste d'avertissements si valeurs hors min/max ou z-score.

    Heuristique simple pour alerter l'utilisateur sans bloquer la prediction.
    """

    warnings = []
    for col, value in inputs.items():
        if col not in profile.numeric_stats:
            continue
        try:
            val = float(value)
        except (TypeError, ValueError):
            continue
        stats = profile.numeric_stats[col]
        if not np.isnan(stats.get("min", np.nan)) and val < stats["min"]:
            warnings.append(f"{col}: {val} < min observe ({stats['min']:.3f})")
        if not np.isnan(stats.get("max", np.nan)) and val > stats["max"]:
            warnings.append(f"{col}: {val} > max observe ({stats['max']:.3f})")
        std = stats.get("std", np.nan)
        mean = stats.get("mean", np.nan)
        if std and not np.isnan(std) and std > 0:
            z = abs((val - mean) / std)
            if z >= 3:
                warnings.append(f"{col}: z-score {z:.2f} (valeur atypique)")
    return warnings


def _percentiles_from_values(values: np.ndarray) -> dict:
    if values.size == 0:
        return {}
    values = values[~np.isnan(values)]
    if values.size == 0:
        return {}
    return {
        "p05": float(np.nanpercentile(values, 5)),
        "p95": float(np.nanpercentile(values, 95)),
        "q1": float(np.nanpercentile(values, 25)),
        "q3": float(np.nanpercentile(values, 75)),
        "iqr": float(np.nanpercentile(values, 75) - np.nanpercentile(values, 25)),
    }


def get_feature_profile(
    profile: DataProfile,
    feature_name: str,
) -> dict:
    """Retourne un profil de feature enrichi (min/max, p05/p95, mean/std)."""
    if not profile:
        return {}

    base = dict(profile.numeric_stats.get(feature_name, {}))

    # Enrichit avec les percentiles si absents.
    if ("p05" not in base or "p95" not in base) and profile.df is not None:
        if feature_name in profile.df.columns:
            values = pd.to_numeric(profile.df[feature_name], errors="coerce").to_numpy()
            base.update(_percentiles_from_values(values))

    if ("p05" not in base or "p95" not in base) and profile.sample_values:
        values = np.asarray(profile.sample_values.get(feature_name, []), dtype=float)
        base.update(_percentiles_from_values(values))

    return base


def _fallback_percentiles(profile: dict) -> tuple[float | None, float | None]:
    """Renvoie p05/p95 avec fallback mean +/- 2*std ou min/max."""
    p05 = profile.get("p05")
    p95 = profile.get("p95")
    if p05 is None or p95 is None or np.isnan(p05) or np.isnan(p95):
        mean = profile.get("mean")
        std = profile.get("std")
        if mean is not None and std not in (None, 0) and not np.isnan(std):
            return float(mean - 2 * std), float(mean + 2 * std)
        min_val = profile.get("min")
        max_val = profile.get("max")
        if min_val is not None and max_val is not None:
            return float(min_val), float(max_val)
        return None, None
    return float(p05), float(p95)


def ood_level(value: float | None, profile: dict) -> dict:
    """Retourne un niveau OOD (ok/warn/out/unknown) et un message."""
    if value is None:
        return {"level": "unknown", "message": "Profil insuffisant", "bounds": {}}
    try:
        val = float(value)
    except (TypeError, ValueError):
        return {"level": "unknown", "message": "Valeur invalide", "bounds": {}}

    min_val = profile.get("min")
    max_val = profile.get("max")
    if min_val is None or max_val is None or np.isnan(min_val) or np.isnan(max_val):
        return {"level": "unknown", "message": "Profil insuffisant", "bounds": {}}

    p05, p95 = _fallback_percentiles(profile)
    bounds = {"min": min_val, "max": max_val, "p05": p05, "p95": p95}

    if val < min_val or val > max_val:
        return {"level": "out", "message": "Hors domaine", "bounds": bounds}
    if p05 is not None and p95 is not None and (val < p05 or val > p95):
        return {"level": "warn", "message": "Attention", "bounds": bounds}
    return {"level": "ok", "message": "OK", "bounds": bounds}


def format_bounds(profile: dict) -> str:
    """Formate min/max et p05/p95 pour affichage UI."""
    min_val = profile.get("min")
    max_val = profile.get("max")
    p05, p95 = _fallback_percentiles(profile)

    parts = []
    if p05 is not None and p95 is not None:
        parts.append(f"p05-p95: {p05:.3f}–{p95:.3f}")
    if min_val is not None and max_val is not None and not np.isnan(min_val) and not np.isnan(max_val):
        parts.append(f"min-max: {min_val:.3f}–{max_val:.3f}")
    return " | ".join(parts)
